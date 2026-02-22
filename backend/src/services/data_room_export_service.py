"""Data Room Export Service — Builds multi-tab XLSX files for org data.

Exports 10 datasets as tabs in a single XLSX workbook.
Filters by org_id, date range, and other parameters.
Strips sensitive columns and truncates oversized JSON values.
"""
import io
import json
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from uuid import UUID

from sqlalchemy.orm import Session

from backend.src.database.models import (
    FlywheelRun,
    FlywheelStep,
    DecisionPack,
    DecisionOutcome,
    DecisionRanking,
    JobRun,
    JobRunStatus,
    MetaAlert,
    Creative,
    MetaInsightsDaily,
)
from src.utils.logging_config import logger


# ── Dataset registry ─────────────────────────────────────────────────────────

# Each entry: (display_name, model_or_None, description)
_DATASET_REGISTRY = {
    "flywheel_runs": ("FlywheelRuns", FlywheelRun, "All flywheel orchestration runs"),
    "flywheel_steps": ("FlywheelSteps", FlywheelStep, "Individual flywheel step results"),
    "decision_queue": ("DecisionQueue", DecisionPack, "Decision packs (queue)"),
    "decision_outcomes": ("DecisionOutcomes", DecisionOutcome, "Decision outcome measurements"),
    "decision_rankings": ("DecisionRankings", DecisionRanking, "Decision ranking scores"),
    "job_runs": ("JobRuns", JobRun, "Background job execution ledger"),
    "alerts": ("Alerts", MetaAlert, "Meta performance alerts"),
    "opportunities": ("Opportunities", None, "Identified market opportunities (from latest analysis job)"),
    "creatives": ("Creatives", Creative, "Generated and tracked creatives"),
    "meta_insights_daily": ("MetaInsightsDaily", MetaInsightsDaily, "Daily Meta Ads performance metrics"),
}


class DataRoomExportService:
    """Build multi-tab XLSX exports with security filtering."""

    SENSITIVE_COLUMNS = {
        "access_token_encrypted",
        "refresh_token_encrypted",
        "password_hash",
        "key_hash",
        "key_prefix",
    }
    MAX_JSON_CELL_LEN = 32768

    def __init__(self, db: Session, org_id):
        self.db = db
        self.org_id = UUID(str(org_id)) if not isinstance(org_id, UUID) else org_id

    # ── Public API ────────────────────────────────────────────────────────────

    def get_schema(self) -> List[Dict[str, Any]]:
        """Return available datasets and their metadata."""
        datasets = []
        for key, (display, model, desc) in _DATASET_REGISTRY.items():
            columns = []
            if model is not None:
                for col in model.__table__.columns:
                    if col.name not in self.SENSITIVE_COLUMNS:
                        columns.append(col.name)
            else:
                # Opportunities — dynamic columns from job result
                columns = [
                    "id", "gap_id", "title", "description", "strategy",
                    "priority", "estimated_impact", "impact_reasoning", "identified_at",
                ]

            datasets.append({
                "key": key,
                "display_name": display,
                "description": desc,
                "columns": columns,
            })
        return datasets

    def build_xlsx(self, params: Dict[str, Any]) -> Tuple[bytes, int]:
        """Build XLSX bytes from parameters. Returns (xlsx_bytes, total_rows)."""
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        # Remove default sheet — we'll create named ones
        default_ws = wb.active

        requested_datasets = params.get("datasets") or list(_DATASET_REGISTRY.keys())
        total_rows = 0
        sheets_created = 0

        for key in requested_datasets:
            if key not in _DATASET_REGISTRY:
                logger.warning("DATA_ROOM_UNKNOWN_DATASET | key={}", key)
                continue

            display_name, model, _desc = _DATASET_REGISTRY[key]

            try:
                rows = self._query_dataset(key, params)
            except Exception as e:
                logger.error("DATA_ROOM_QUERY_ERROR | dataset={} | error={}", key, str(e)[:200])
                continue

            # Create sheet (max 31 chars for Excel sheet name)
            sheet_name = display_name[:31]
            if sheets_created == 0:
                ws = default_ws
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            sheets_created += 1

            # Determine columns: from data if available, else from model schema
            if rows and isinstance(rows[0], dict):
                col_names = list(rows[0].keys())
            elif rows:
                col_names = [
                    c.name for c in rows[0].__class__.__table__.columns
                    if c.name not in self.SENSITIVE_COLUMNS
                ]
            elif model is not None:
                col_names = [
                    c.name for c in model.__table__.columns
                    if c.name not in self.SENSITIVE_COLUMNS
                ]
            else:
                # Opportunities (dynamic schema)
                col_names = [
                    "id", "gap_id", "title", "description", "strategy",
                    "priority", "estimated_impact", "impact_reasoning", "identified_at",
                ]

            # Write header
            ws.append(col_names)

            # Write data rows
            for row_obj in rows:
                row_values = []
                for col_name in col_names:
                    if isinstance(row_obj, dict):
                        raw = row_obj.get(col_name, "")
                    else:
                        raw = getattr(row_obj, col_name, "")
                    row_values.append(self._sanitize_value(raw, col_name))
                ws.append(row_values)
                total_rows += 1

        # If no data at all, write an info sheet
        if sheets_created == 0:
            ws = default_ws
            ws.title = "Info"
            ws.append(["No data available for the selected filters."])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        logger.info(
            "DATA_ROOM_XLSX_BUILT | org={} | datasets={} | total_rows={}",
            self.org_id, sheets_created, total_rows,
        )
        return buf.getvalue(), total_rows

    # ── Private helpers ──────────────────────────────────────────────────────

    def _query_dataset(self, key: str, params: Dict[str, Any]) -> List:
        """Query a single dataset with filters applied."""
        limit = params.get("limit", 10000)
        date_from = params.get("date_from")
        date_to = params.get("date_to")
        status_filter = params.get("status")
        ad_account_id = params.get("ad_account_id")
        entity_type = params.get("entity_type")
        severity = params.get("severity")

        # Parse date strings if provided
        if isinstance(date_from, str):
            try:
                date_from = datetime.fromisoformat(date_from)
            except (ValueError, TypeError):
                date_from = None
        if isinstance(date_to, str):
            try:
                date_to = datetime.fromisoformat(date_to)
            except (ValueError, TypeError):
                date_to = None

        # Special case: Opportunities (not a DB model)
        if key == "opportunities":
            return self._query_opportunities(limit)

        _display, model, _desc = _DATASET_REGISTRY[key]
        if model is None:
            return []

        query = self.db.query(model)

        # Filter by org_id (all models have org_id except flywheel_steps)
        if hasattr(model, "org_id"):
            query = query.filter(model.org_id == self.org_id)
        elif key == "flywheel_steps":
            # Join through flywheel_runs to filter by org
            query = query.join(FlywheelRun, FlywheelStep.flywheel_run_id == FlywheelRun.id)
            query = query.filter(FlywheelRun.org_id == self.org_id)

        # Date filters (use created_at if available)
        if date_from and hasattr(model, "created_at"):
            query = query.filter(model.created_at >= date_from)
        if date_to and hasattr(model, "created_at"):
            query = query.filter(model.created_at <= date_to)

        # Status filter
        if status_filter and hasattr(model, "status"):
            query = query.filter(model.status == status_filter)

        # Ad account filter
        if ad_account_id and hasattr(model, "ad_account_id"):
            query = query.filter(model.ad_account_id == UUID(ad_account_id))

        # Entity type filter
        if entity_type and hasattr(model, "entity_type"):
            query = query.filter(model.entity_type == entity_type)

        # Severity filter (for alerts)
        if severity and hasattr(model, "severity"):
            query = query.filter(model.severity == severity)

        # Order by created_at descending if available
        if hasattr(model, "created_at"):
            query = query.order_by(model.created_at.desc())

        return query.limit(limit).all()

    def _query_opportunities(self, limit: int) -> List[Dict[str, Any]]:
        """Get opportunities from the latest successful analysis job."""
        job = (
            self.db.query(JobRun)
            .filter(
                JobRun.org_id == self.org_id,
                JobRun.job_type.in_(["opportunities_analyze", "unified_intelligence_analyze"]),
                JobRun.status == JobRunStatus.SUCCEEDED,
            )
            .order_by(JobRun.finished_at.desc())
            .first()
        )

        if not job or not job.payload_json:
            return []

        result = job.payload_json.get("result", [])
        if isinstance(result, list):
            return result[:limit]
        return []

    def _sanitize_value(self, value: Any, col_name: str) -> Any:
        """Remove sensitive data, truncate JSON, convert to cell-safe types."""
        # Never include sensitive columns
        if col_name in self.SENSITIVE_COLUMNS:
            return "[REDACTED]"

        if value is None:
            return ""

        # UUID → string
        if isinstance(value, UUID):
            return str(value)

        # datetime → ISO string
        if isinstance(value, datetime):
            return value.isoformat()

        # Enum → value
        if hasattr(value, "value"):
            return value.value

        # dict / list → JSON string (truncated)
        if isinstance(value, (dict, list)):
            try:
                json_str = json.dumps(value, default=str, ensure_ascii=False)
                if len(json_str) > self.MAX_JSON_CELL_LEN:
                    return json_str[: self.MAX_JSON_CELL_LEN] + "...[TRUNCATED]"
                return json_str
            except (TypeError, ValueError):
                return str(value)[:self.MAX_JSON_CELL_LEN]

        # String truncation for very long strings
        if isinstance(value, str) and len(value) > self.MAX_JSON_CELL_LEN:
            return value[: self.MAX_JSON_CELL_LEN] + "...[TRUNCATED]"

        return value
