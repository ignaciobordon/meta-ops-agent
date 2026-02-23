"""
Sprint 5 – BLOQUE F: Brain Stats API
Operator Brain dashboard data — top features, recent outcomes, entity trust.
When learning tables are empty, derives intelligence from real Meta campaign data.
"""
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional
from uuid import UUID
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from backend.src.database.models import (
    DecisionOutcome,
    EntityMemory,
    FeatureMemory,
    InsightLevel,
    MetaCampaign,
    MetaInsightsDaily,
    OutcomeLabel,
)
from loguru import logger

from backend.src.database.session import get_db
from backend.src.middleware.auth import get_current_user

router = APIRouter()

# ── Human-readable objective names ────────────────────────────────────────────

_OBJECTIVE_LABELS = {
    "OUTCOME_LEADS": "Leads",
    "OUTCOME_ENGAGEMENT": "Engagement",
    "OUTCOME_AWARENESS": "Awareness",
    "OUTCOME_TRAFFIC": "Traffic",
    "OUTCOME_SALES": "Sales",
    "OUTCOME_APP_PROMOTION": "App Promotion",
    "CONVERSIONS": "Conversions",
    "LINK_CLICKS": "Link Clicks",
    "REACH": "Reach",
    "BRAND_AWARENESS": "Brand Awareness",
    "POST_ENGAGEMENT": "Post Engagement",
    "VIDEO_VIEWS": "Video Views",
    "MESSAGES": "Messages",
    "STORE_VISITS": "Store Visits",
}


def _clean_objective(raw: str) -> str:
    if raw in _OBJECTIVE_LABELS:
        return _OBJECTIVE_LABELS[raw]
    return raw.replace("OUTCOME_", "").replace("_", " ").title()


def _parse_dates(days: int, since: Optional[str], until: Optional[str]):
    """Parse date range from days offset or explicit since/until."""
    if since:
        since_dt = datetime.strptime(since, "%Y-%m-%d")
    else:
        since_dt = datetime.utcnow() - timedelta(days=days)

    until_dt = None
    if until:
        until_dt = datetime.strptime(until, "%Y-%m-%d") + timedelta(days=1)

    return since_dt, until_dt


# ── Schemas ───────────────────────────────────────────────────────────────────


class FeatureStats(BaseModel):
    feature_type: str
    feature_key: str
    win_rate: float
    samples: int
    avg_delta: Dict[str, Any] = {}


class RecentOutcome(BaseModel):
    entity_type: str
    entity_id: str
    action_type: str
    outcome_label: str
    confidence: float
    horizon_minutes: int
    executed_at: datetime
    detail: Dict[str, Any] = {}


class EntityTrust(BaseModel):
    entity_type: str
    entity_id: str
    trust_score: float
    last_outcome: Optional[str] = None
    last_seen_at: Optional[datetime] = None
    detail: Dict[str, Any] = {}


class BrainSummary(BaseModel):
    total_campaigns: int = 0
    avg_trust: float = 0
    win_count: int = 0
    loss_count: int = 0
    total_spend: float = 0
    total_clicks: int = 0
    total_impressions: int = 0
    avg_ctr: float = 0
    avg_cpc: float = 0
    avg_cpm: float = 0
    period_days: int = 90
    # Trend vs previous period
    spend_trend: Optional[float] = None
    ctr_trend: Optional[float] = None
    cpc_trend: Optional[float] = None


class BrainStatsResponse(BaseModel):
    summary: BrainSummary
    top_features: List[FeatureStats]
    recent_outcomes: List[RecentOutcome]
    entity_trust: List[EntityTrust]


# ── Helpers: derive brain data from Meta insights ─────────────────────────────


def _compute_summary(
    db: Session, org_uuid: UUID, since_dt: datetime, until_dt: Optional[datetime], days: int,
) -> BrainSummary:
    """Compute overall intelligence summary with trend comparison."""
    query = db.query(
        func.sum(MetaInsightsDaily.spend).label("spend"),
        func.sum(MetaInsightsDaily.impressions).label("impressions"),
        func.sum(MetaInsightsDaily.clicks).label("clicks"),
        func.count(func.distinct(MetaInsightsDaily.entity_meta_id)).label("campaigns"),
    ).filter(
        MetaInsightsDaily.org_id == org_uuid,
        MetaInsightsDaily.level == InsightLevel.CAMPAIGN,
        MetaInsightsDaily.date_start >= since_dt,
    )
    if until_dt:
        query = query.filter(MetaInsightsDaily.date_start < until_dt)
    cur = query.first()

    spend = float(cur.spend or 0)
    impressions = int(cur.impressions or 0)
    clicks = int(cur.clicks or 0)
    campaigns = int(cur.campaigns or 0)
    ctr = (clicks / impressions * 100) if impressions > 0 else 0
    cpc = (spend / clicks) if clicks > 0 else 0
    cpm = (spend / impressions * 1000) if impressions > 0 else 0

    # Previous period for trend
    period_len = (until_dt - since_dt).days if until_dt else days
    prev_since = since_dt - timedelta(days=period_len)
    prev_until = since_dt

    prev_q = db.query(
        func.sum(MetaInsightsDaily.spend).label("spend"),
        func.sum(MetaInsightsDaily.impressions).label("impressions"),
        func.sum(MetaInsightsDaily.clicks).label("clicks"),
    ).filter(
        MetaInsightsDaily.org_id == org_uuid,
        MetaInsightsDaily.level == InsightLevel.CAMPAIGN,
        MetaInsightsDaily.date_start >= prev_since,
        MetaInsightsDaily.date_start < prev_until,
    ).first()

    prev_spend = float(prev_q.spend or 0)
    prev_impressions = int(prev_q.impressions or 0)
    prev_clicks = int(prev_q.clicks or 0)
    prev_ctr = (prev_clicks / prev_impressions * 100) if prev_impressions > 0 else 0
    prev_cpc = (prev_spend / prev_clicks) if prev_clicks > 0 else 0

    def pct_change(cur_val, prev_val):
        if prev_val == 0:
            return None
        return round((cur_val - prev_val) / prev_val * 100, 1)

    return BrainSummary(
        total_campaigns=campaigns,
        total_spend=round(spend, 2),
        total_clicks=clicks,
        total_impressions=impressions,
        avg_ctr=round(ctr, 2),
        avg_cpc=round(cpc, 2),
        avg_cpm=round(cpm, 2),
        period_days=period_len,
        spend_trend=pct_change(spend, prev_spend),
        ctr_trend=pct_change(ctr, prev_ctr),
        cpc_trend=pct_change(cpc, prev_cpc),
    )


def _derive_entity_trust(
    db: Session, org_uuid: UUID, since_dt: datetime, until_dt: Optional[datetime],
) -> List[EntityTrust]:
    """Derive entity trust from campaign performance (CTR, spend efficiency)."""
    query = db.query(
        MetaInsightsDaily.entity_meta_id,
        func.sum(MetaInsightsDaily.spend).label("total_spend"),
        func.sum(MetaInsightsDaily.impressions).label("total_impressions"),
        func.sum(MetaInsightsDaily.clicks).label("total_clicks"),
        func.sum(MetaInsightsDaily.conversions).label("total_conversions"),
        func.count().label("day_count"),
        func.max(MetaInsightsDaily.date_start).label("last_date"),
    ).filter(
        MetaInsightsDaily.org_id == org_uuid,
        MetaInsightsDaily.level == InsightLevel.CAMPAIGN,
        MetaInsightsDaily.date_start >= since_dt,
    )
    if until_dt:
        query = query.filter(MetaInsightsDaily.date_start < until_dt)

    rows = query.group_by(
        MetaInsightsDaily.entity_meta_id,
    ).having(
        func.sum(MetaInsightsDaily.spend) > 0,
    ).order_by(
        func.sum(MetaInsightsDaily.spend).desc(),
    ).limit(50).all()

    trust_entries = []
    for r in rows:
        spend = float(r.total_spend or 0)
        impressions = int(r.total_impressions or 0)
        clicks = int(r.total_clicks or 0)
        conversions = int(r.total_conversions or 0)
        ctr = (clicks / impressions * 100) if impressions > 0 else 0
        cpc = (spend / clicks) if clicks > 0 else 999
        cpm = (spend / impressions * 1000) if impressions > 0 else 0

        # Trust score: weighted formula
        ctr_score = min(ctr * 20, 40)
        efficiency_score = max(0, 30 - cpc * 2)
        stability_score = min(r.day_count * 2, 30)
        trust = min(round(ctr_score + efficiency_score + stability_score, 1), 100)

        campaign = db.query(MetaCampaign).filter(
            MetaCampaign.org_id == org_uuid,
            MetaCampaign.meta_campaign_id == r.entity_meta_id,
        ).first()

        label = "win" if ctr > 1.0 else "neutral" if ctr > 0.5 else "loss"
        objective = _clean_objective(campaign.objective) if campaign and campaign.objective else None

        trust_entries.append(EntityTrust(
            entity_type="campaign",
            entity_id=campaign.name[:40] if campaign and campaign.name else r.entity_meta_id,
            trust_score=trust,
            last_outcome=label,
            last_seen_at=r.last_date,
            detail={
                "spend": round(spend, 2),
                "clicks": clicks,
                "impressions": impressions,
                "conversions": conversions,
                "ctr": round(ctr, 2),
                "cpc": round(cpc, 2),
                "cpm": round(cpm, 2),
                "days_active": r.day_count,
                "objective": objective,
                "status": campaign.effective_status if campaign else None,
                "entity_meta_id": r.entity_meta_id,
                "trust_ctr_score": round(ctr_score, 1),
                "trust_efficiency_score": round(efficiency_score, 1),
                "trust_stability_score": round(stability_score, 1),
                "trust_formula": "CTR\u00d720 (max 40) + Efficiency (max 30) + Stability (max 30)",
            },
        ))

    return trust_entries


def _derive_top_features(
    db: Session, org_uuid: UUID, since_dt: datetime, until_dt: Optional[datetime],
) -> List[FeatureStats]:
    """Derive feature performance from campaign objectives and statuses."""
    campaigns = db.query(MetaCampaign).filter(
        MetaCampaign.org_id == org_uuid,
        MetaCampaign.objective.isnot(None),
    ).all()

    objective_stats: Dict[str, Dict] = {}
    for c in campaigns:
        obj = c.objective or "unknown"
        if obj not in objective_stats:
            objective_stats[obj] = {
                "wins": 0, "total": 0, "spend": 0, "clicks": 0,
                "impressions": 0, "conversions": 0, "campaigns": [],
            }

        query = db.query(
            func.sum(MetaInsightsDaily.spend).label("spend"),
            func.sum(MetaInsightsDaily.clicks).label("clicks"),
            func.sum(MetaInsightsDaily.impressions).label("impressions"),
            func.sum(MetaInsightsDaily.conversions).label("conversions"),
        ).filter(
            MetaInsightsDaily.org_id == org_uuid,
            MetaInsightsDaily.entity_meta_id == c.meta_campaign_id,
            MetaInsightsDaily.date_start >= since_dt,
        )
        if until_dt:
            query = query.filter(MetaInsightsDaily.date_start < until_dt)
        perf = query.first()

        if perf and perf.spend:
            clicks = int(perf.clicks or 0)
            impressions = int(perf.impressions or 0)
            conversions = int(perf.conversions or 0)
            ctr = (clicks / impressions * 100) if impressions > 0 else 0
            objective_stats[obj]["total"] += 1
            objective_stats[obj]["spend"] += float(perf.spend or 0)
            objective_stats[obj]["clicks"] += clicks
            objective_stats[obj]["impressions"] += impressions
            objective_stats[obj]["conversions"] += conversions
            objective_stats[obj]["campaigns"].append(c.name or c.meta_campaign_id)
            if ctr > 1.0:
                objective_stats[obj]["wins"] += 1

    # Also derive by campaign status
    status_stats: Dict[str, Dict] = {}
    for c in campaigns:
        status = c.effective_status or c.status or "unknown"
        if status not in status_stats:
            status_stats[status] = {"wins": 0, "total": 0, "spend": 0, "clicks": 0, "impressions": 0}

        query = db.query(
            func.sum(MetaInsightsDaily.spend).label("spend"),
            func.sum(MetaInsightsDaily.clicks).label("clicks"),
            func.sum(MetaInsightsDaily.impressions).label("impressions"),
        ).filter(
            MetaInsightsDaily.org_id == org_uuid,
            MetaInsightsDaily.entity_meta_id == c.meta_campaign_id,
            MetaInsightsDaily.date_start >= since_dt,
        )
        if until_dt:
            query = query.filter(MetaInsightsDaily.date_start < until_dt)
        perf = query.first()

        if perf and perf.spend:
            clicks = int(perf.clicks or 0)
            impressions = int(perf.impressions or 0)
            ctr = (clicks / impressions * 100) if impressions > 0 else 0
            status_stats[status]["total"] += 1
            status_stats[status]["spend"] += float(perf.spend or 0)
            status_stats[status]["clicks"] += clicks
            status_stats[status]["impressions"] += impressions
            if ctr > 1.0:
                status_stats[status]["wins"] += 1

    features = []

    # Objective features
    for obj, stats in objective_stats.items():
        if stats["total"] >= 1:
            avg_ctr = (stats["clicks"] / stats["impressions"] * 100) if stats["impressions"] > 0 else 0
            avg_cpc = (stats["spend"] / stats["clicks"]) if stats["clicks"] > 0 else 0
            cpm = (stats["spend"] / stats["impressions"] * 1000) if stats["impressions"] > 0 else 0
            features.append(FeatureStats(
                feature_type="Objetivo",
                feature_key=_clean_objective(obj),
                win_rate=stats["wins"] / stats["total"] if stats["total"] > 0 else 0,
                samples=stats["total"],
                avg_delta={
                    "spend": round(stats["spend"], 2),
                    "clicks": stats["clicks"],
                    "impressions": stats["impressions"],
                    "conversions": stats["conversions"],
                    "avg_ctr": round(avg_ctr, 2),
                    "avg_cpc": round(avg_cpc, 2),
                    "cpm": round(cpm, 2),
                    "campaigns": stats["campaigns"][:5],
                },
            ))

    # Status features
    for status, stats in status_stats.items():
        if stats["total"] >= 1:
            avg_ctr = (stats["clicks"] / stats["impressions"] * 100) if stats["impressions"] > 0 else 0
            features.append(FeatureStats(
                feature_type="Estado",
                feature_key=status.replace("_", " ").title(),
                win_rate=stats["wins"] / stats["total"] if stats["total"] > 0 else 0,
                samples=stats["total"],
                avg_delta={
                    "spend": round(stats["spend"], 2),
                    "clicks": stats["clicks"],
                    "avg_ctr": round(avg_ctr, 2),
                },
            ))

    features.sort(key=lambda f: f.win_rate, reverse=True)
    return features[:20]


def _derive_recent_outcomes(
    db: Session, org_uuid: UUID, since_dt: datetime, until_dt: Optional[datetime],
) -> List[RecentOutcome]:
    """Derive entity-level outcomes from campaign performance over the period.

    Instead of noisy day-by-day rows, aggregates per campaign:
    - Groups by entity_meta_id over the selected period
    - Compares to previous period for trend delta
    - Labels using composite signal (CTR + trend + conversions)
    - Confidence based on volume (impressions/1000)
    """
    # Current period aggregates
    query = db.query(
        MetaInsightsDaily.entity_meta_id,
        func.sum(MetaInsightsDaily.spend).label("total_spend"),
        func.sum(MetaInsightsDaily.clicks).label("total_clicks"),
        func.sum(MetaInsightsDaily.impressions).label("total_impressions"),
        func.sum(MetaInsightsDaily.conversions).label("total_conversions"),
        func.count().label("day_count"),
        func.max(MetaInsightsDaily.date_start).label("last_date"),
    ).filter(
        MetaInsightsDaily.org_id == org_uuid,
        MetaInsightsDaily.level == InsightLevel.CAMPAIGN,
        MetaInsightsDaily.date_start >= since_dt,
        MetaInsightsDaily.spend > 0,
    )
    if until_dt:
        query = query.filter(MetaInsightsDaily.date_start < until_dt)

    rows = query.group_by(
        MetaInsightsDaily.entity_meta_id,
    ).having(
        func.sum(MetaInsightsDaily.spend) > 0,
    ).order_by(
        func.sum(MetaInsightsDaily.spend).desc(),
    ).limit(50).all()

    if not rows:
        return []

    # Previous period for trend comparison
    period_len = (until_dt - since_dt).days if until_dt else (datetime.utcnow() - since_dt).days
    prev_since = since_dt - timedelta(days=max(period_len, 1))
    prev_until = since_dt

    prev_query = db.query(
        MetaInsightsDaily.entity_meta_id,
        func.sum(MetaInsightsDaily.clicks).label("prev_clicks"),
        func.sum(MetaInsightsDaily.impressions).label("prev_impressions"),
    ).filter(
        MetaInsightsDaily.org_id == org_uuid,
        MetaInsightsDaily.level == InsightLevel.CAMPAIGN,
        MetaInsightsDaily.date_start >= prev_since,
        MetaInsightsDaily.date_start < prev_until,
        MetaInsightsDaily.spend > 0,
    ).group_by(MetaInsightsDaily.entity_meta_id).all()

    prev_map = {r.entity_meta_id: r for r in prev_query}

    # Campaign name cache
    campaign_ids = [r.entity_meta_id for r in rows]
    campaigns = db.query(MetaCampaign).filter(
        MetaCampaign.org_id == org_uuid,
        MetaCampaign.meta_campaign_id.in_(campaign_ids),
    ).all()
    campaigns_map = {c.meta_campaign_id: c for c in campaigns}

    outcomes = []
    for r in rows:
        spend = float(r.total_spend or 0)
        clicks = int(r.total_clicks or 0)
        impressions = int(r.total_impressions or 0)
        conversions = int(r.total_conversions or 0)
        ctr = (clicks / impressions * 100) if impressions > 0 else 0
        cpc = (spend / clicks) if clicks > 0 else 0
        cpm = (spend / impressions * 1000) if impressions > 0 else 0

        # Trend: compare CTR to previous period
        prev = prev_map.get(r.entity_meta_id)
        trend_improving = False
        trend_declining = False
        if prev:
            prev_clicks = int(prev.prev_clicks or 0)
            prev_impressions = int(prev.prev_impressions or 0)
            prev_ctr = (prev_clicks / prev_impressions * 100) if prev_impressions > 0 else 0
            if prev_ctr > 0:
                trend_improving = ctr > prev_ctr
                trend_declining = ctr < prev_ctr * 0.9

        # Label using composite signal
        if ctr > 1.0 and (trend_improving or conversions > 0):
            label = "win"
        elif ctr < 0.5 and trend_declining:
            label = "loss"
        else:
            label = "neutral"

        # Confidence based on volume (statistical significance)
        confidence = min(impressions / 1000.0, 1.0)

        campaign = campaigns_map.get(r.entity_meta_id)

        outcomes.append(RecentOutcome(
            entity_type="campaign",
            entity_id=campaign.name[:40] if campaign and campaign.name else r.entity_meta_id,
            action_type="period_performance",
            outcome_label=label,
            confidence=round(confidence, 2),
            horizon_minutes=period_len * 1440,
            executed_at=r.last_date,
            detail={
                "spend": round(spend, 2),
                "clicks": clicks,
                "impressions": impressions,
                "conversions": conversions,
                "ctr": round(ctr, 2),
                "cpc": round(cpc, 2),
                "cpm": round(cpm, 2),
                "days_active": r.day_count,
                "trend_improving": trend_improving,
            },
        ))

    return outcomes


# ── Endpoints ─────────────────────────────────────────────────────────────────


@router.get("/brain/stats", response_model=BrainStatsResponse)
def get_brain_stats(
    days: int = Query(90, le=730, description="Lookback window in days"),
    since: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    until: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get Operator Brain dashboard data for the user's organization."""
    org_id = user.get("org_id", "")

    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)

    # Try learning tables first
    features = db.query(FeatureMemory).filter(
        FeatureMemory.org_id == org_uuid,
        FeatureMemory.samples >= 3,
    ).order_by(FeatureMemory.win_rate.desc()).limit(20).all()

    outcomes = db.query(DecisionOutcome).filter(
        DecisionOutcome.org_id == org_uuid,
        DecisionOutcome.outcome_label != OutcomeLabel.UNKNOWN,
    ).order_by(DecisionOutcome.created_at.desc()).limit(20).all()

    entities = db.query(EntityMemory).filter(
        EntityMemory.org_id == org_uuid,
    ).order_by(EntityMemory.last_seen_at.desc()).limit(20).all()

    has_learning_data = bool(features or outcomes or entities)

    # Always compute summary from Meta insights
    summary = _compute_summary(db, org_uuid, since_dt, until_dt, days)

    # Always derive entities from MetaInsightsDaily for full detail + trust breakdown
    entity_trust = _derive_entity_trust(db, org_uuid, since_dt, until_dt)

    # If EntityMemory has learned trust scores, overlay them (more accurate than formula)
    # and add any EntityMemory entries not found in derived data
    if entities:
        derived_ids = {et.entity_id for et in entity_trust}
        derived_meta_ids = {(et.detail or {}).get("entity_meta_id", "") for et in entity_trust}
        memory_map: Dict[str, EntityMemory] = {}
        for em in entities:
            memory_map[em.entity_id] = em
        for et in entity_trust:
            d = et.detail or {}
            meta_id = d.get("entity_meta_id", "")
            mem = memory_map.get(et.entity_id) or memory_map.get(meta_id)
            if mem:
                et.trust_score = mem.trust_score
                et.last_outcome = mem.last_outcome_label.value if mem.last_outcome_label else et.last_outcome
                et.last_seen_at = mem.last_seen_at or et.last_seen_at
        # Add EntityMemory entries missing from MetaInsightsDaily (e.g. entities
        # tracked by the learning loop but with no daily insights yet)
        for em in entities:
            if em.entity_id not in derived_ids and em.entity_id not in derived_meta_ids:
                entity_trust.append(EntityTrust(
                    entity_type=em.entity_type,
                    entity_id=em.entity_id,
                    trust_score=em.trust_score,
                    last_outcome=em.last_outcome_label.value if em.last_outcome_label else None,
                    last_seen_at=em.last_seen_at,
                ))

    if has_learning_data:
        top_features = [
            FeatureStats(
                feature_type=f.feature_type.value if hasattr(f.feature_type, 'value') else str(f.feature_type),
                feature_key=f.feature_key,
                win_rate=f.win_rate,
                samples=f.samples,
                avg_delta=f.avg_delta_json or {},
            )
            for f in features
        ]

        recent_outcomes = [
            RecentOutcome(
                entity_type=o.entity_type,
                entity_id=o.entity_id,
                action_type=o.action_type.value if hasattr(o.action_type, 'value') else str(o.action_type),
                outcome_label=o.outcome_label.value if hasattr(o.outcome_label, 'value') else str(o.outcome_label),
                confidence=o.confidence,
                horizon_minutes=o.horizon_minutes,
                executed_at=o.executed_at,
            )
            for o in outcomes
        ]
    else:
        top_features = _derive_top_features(db, org_uuid, since_dt, until_dt)
        recent_outcomes = _derive_recent_outcomes(db, org_uuid, since_dt, until_dt)

    # Fill summary from trust data
    summary.total_campaigns = len(entity_trust)
    summary.win_count = sum(1 for e in entity_trust if e.last_outcome == "win")
    summary.loss_count = sum(1 for e in entity_trust if e.last_outcome == "loss")
    summary.avg_trust = round(
        sum(e.trust_score for e in entity_trust) / len(entity_trust), 1
    ) if entity_trust else 0

    return BrainStatsResponse(
        summary=summary,
        top_features=top_features,
        recent_outcomes=recent_outcomes,
        entity_trust=entity_trust,
    )


# ── Suggestions Endpoint ─────────────────────────────────────────────────────


class BrainSuggestion(BaseModel):
    type: str  # scale, pause, test, refresh, optimize
    entity_id: str
    title: str
    description: str
    metrics: Dict[str, Any] = {}


@router.get("/brain/suggestions")
def get_brain_suggestions(
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get rule-based actionable suggestions from Brain data."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)
    entity_trust = _derive_entity_trust(db, org_uuid, since_dt, until_dt)
    top_features = _derive_top_features(db, org_uuid, since_dt, until_dt)

    suggestions: List[BrainSuggestion] = []

    for e in entity_trust:
        d = e.detail or {}
        ctr = d.get("ctr", 0)
        cpc = d.get("cpc", 0)
        frequency = d.get("frequency", 0)
        spend = d.get("spend", 0)
        days_active = d.get("days_active", 0)

        # High trust + high CTR → Scale
        if e.trust_score > 70 and ctr > 1.5:
            suggestions.append(BrainSuggestion(
                type="scale",
                entity_id=e.entity_id,
                title=f"Scale: Increase budget for {e.entity_id}",
                description=f"Trust score {e.trust_score:.0f}/100 with strong CTR ({ctr}%). This campaign shows consistent high performance — consider increasing daily budget by 20-30%.",
                metrics={"trust": e.trust_score, "ctr": ctr, "spend": spend},
            ))

        # Low trust + low CTR → Pause
        elif e.trust_score < 30 and ctr < 0.5:
            suggestions.append(BrainSuggestion(
                type="pause",
                entity_id=e.entity_id,
                title=f"Pause: {e.entity_id} underperforming",
                description=f"Trust score {e.trust_score:.0f}/100 with very low CTR ({ctr}%). Consider pausing this campaign and reallocating budget to better performers.",
                metrics={"trust": e.trust_score, "ctr": ctr, "spend": spend},
            ))

        # High CPC → Optimize
        elif cpc > 5 and spend > 100:
            suggestions.append(BrainSuggestion(
                type="optimize",
                entity_id=e.entity_id,
                title=f"Optimize: High CPC on {e.entity_id}",
                description=f"CPC at ${cpc:.2f} is above efficient range. Review audience targeting, bid strategy, or ad relevance to reduce cost per click.",
                metrics={"cpc": cpc, "spend": spend, "ctr": ctr},
            ))

    # Feature-based suggestions
    for f in top_features:
        if f.feature_type == "Objetivo" and f.win_rate < 0.3 and f.samples >= 2:
            suggestions.append(BrainSuggestion(
                type="test",
                entity_id=f.feature_key,
                title=f"Test: New creatives for {f.feature_key} objective",
                description=f"Win rate of {f.win_rate*100:.0f}% across {f.samples} campaigns. The {f.feature_key} objective needs fresh creative angles to improve performance.",
                metrics={"win_rate": round(f.win_rate, 2), "samples": f.samples},
            ))

    # Sort: scale first, then pause, then test, then optimize
    type_order = {"scale": 0, "pause": 1, "refresh": 2, "test": 3, "optimize": 4}
    suggestions.sort(key=lambda s: type_order.get(s.type, 5))

    return {"suggestions": [s.dict() for s in suggestions[:20]]}


# ── Deep Analysis & Flywheel ─────────────────────────────────────────────────

_BRAIN_SYSTEM_PROMPT = """You are a senior performance marketing analyst writing a detailed campaign intelligence report.
Your reports are data-driven, specific, and actionable. Write in professional Spanish.

Rules:
- Reference the EXACT numbers provided — never invent data.
- Compare metrics against standard industry benchmarks for paid social (Meta Ads).
- Be specific: instead of "CTR is low" say "CTR at 1.04% is below the Meta Ads benchmark of 1.5-2.0%."
- Provide 3-5 concrete, prioritized action items with expected impact.
- Structure the report with clear sections using markdown headers.
- Do NOT use emojis or special Unicode symbols. Use plain text only.
"""

_ENTITY_USER_TEMPLATE = """Analyze this advertising campaign entity:

ENTITY: {name}
TRUST SCORE: {trust}/100
 - CTR Component: {ctr_score}/40
 - Efficiency Component: {eff_score}/30
 - Stability Component: {stab_score}/30
PERFORMANCE LABEL: {label}
OBJECTIVE: {objective}
STATUS: {status}

METRICS:
- Spend: ${spend}
- Clicks: {clicks}
- Impressions: {impressions}
- Conversions: {conversions}
- CTR: {ctr}%
- CPC: ${cpc}
- CPM: ${cpm}
- Days Active: {days_active}

Write the report with these sections:
1. **Resumen** — 2-3 sentence overview
2. **Hallazgos Clave** — 3-5 bullet points with specific metrics
3. **Recomendaciones Accionables** — 3-5 specific actions with expected impact
4. **Siguiente Paso en el Flywheel** — Which module to use next: Brand Profile / Opportunities / Market Saturation / Creatives, and why
"""

_FEATURE_USER_TEMPLATE = """Analyze this advertising feature/objective:

FEATURE: {name} (Type: {ftype})
WIN RATE: {win_rate}%
CAMPAIGNS: {samples}

AGGREGATED METRICS:
- Total Spend: ${spend}
- Total Clicks: {clicks}
- Total Impressions: {impressions}
- Conversions: {conversions}
- Average CTR: {avg_ctr}%
- Average CPC: ${avg_cpc}
- CPM: ${cpm}

Write the report with these sections:
1. **Resumen** — 2-3 sentence overview
2. **Hallazgos Clave** — 3-5 bullet points with specific metrics
3. **Mejores Practicas** — 3-4 best practices for this objective type
4. **Recomendaciones Accionables** — 3-5 specific actions with expected impact
5. **Siguiente Paso en el Flywheel** — Which module to use next
"""

_OUTCOME_USER_TEMPLATE = """Analyze this campaign outcome:

CAMPAIGN: {name}
OUTCOME: {label}
CONFIDENCE: {confidence}%

METRICS:
- Spend: ${spend}
- Clicks: {clicks}
- Impressions: {impressions}
- Conversions: {conversions}
- CTR: {ctr}%
- CPC: ${cpc}
- CPM: ${cpm}
- Days Active: {days_active}
- Trend Improving: {trend_improving}

Write the report with these sections:
1. **Resumen** — 2-3 sentence overview of the outcome
2. **Hallazgos Clave** — 3-5 bullet points with specific metrics
3. **Lecciones Aprendidas** — 3-4 key lessons from this outcome
4. **Recomendaciones Accionables** — 3-5 specific next actions
5. **Siguiente Paso en el Flywheel** — Which module to use next
"""


def _generate_deep_analysis(data: dict, analysis_type: str) -> str:
    """Call LLM for deep analysis. Fallback to rule-based if unavailable."""
    try:
        from backend.src.llm.router import get_llm_router
        from backend.src.llm.schema import LLMRequest

        if analysis_type == "entity":
            d = data.get("detail", {})
            user_content = _ENTITY_USER_TEMPLATE.format(
                name=data.get("entity_id", "Unknown"),
                trust=data.get("trust_score", 0),
                ctr_score=d.get("trust_ctr_score", 0),
                eff_score=d.get("trust_efficiency_score", 0),
                stab_score=d.get("trust_stability_score", 0),
                label=data.get("last_outcome", "neutral"),
                objective=d.get("objective", "N/A"),
                status=d.get("status", "N/A"),
                spend=d.get("spend", 0),
                clicks=d.get("clicks", 0),
                impressions=d.get("impressions", 0),
                conversions=d.get("conversions", 0),
                ctr=d.get("ctr", 0),
                cpc=d.get("cpc", 0),
                cpm=d.get("cpm", 0),
                days_active=d.get("days_active", 0),
            )
        elif analysis_type == "feature":
            ad = data.get("avg_delta", {})
            user_content = _FEATURE_USER_TEMPLATE.format(
                name=data.get("feature_key", "Unknown"),
                ftype=data.get("feature_type", ""),
                win_rate=round(data.get("win_rate", 0) * 100),
                samples=data.get("samples", 0),
                spend=ad.get("spend", 0),
                clicks=ad.get("clicks", 0),
                impressions=ad.get("impressions", 0),
                conversions=ad.get("conversions", 0),
                avg_ctr=ad.get("avg_ctr", 0),
                avg_cpc=ad.get("avg_cpc", 0),
                cpm=ad.get("cpm", 0),
            )
        elif analysis_type == "outcome":
            d = data.get("detail", {})
            user_content = _OUTCOME_USER_TEMPLATE.format(
                name=data.get("entity_id", "Unknown"),
                label=data.get("outcome_label", "neutral"),
                confidence=round(data.get("confidence", 0) * 100),
                spend=d.get("spend", 0),
                clicks=d.get("clicks", 0),
                impressions=d.get("impressions", 0),
                conversions=d.get("conversions", 0),
                ctr=d.get("ctr", 0),
                cpc=d.get("cpc", 0),
                cpm=d.get("cpm", 0),
                days_active=d.get("days_active", 0),
                trend_improving=d.get("trend_improving", False),
            )
        else:
            return ""

        request = LLMRequest(
            task_type="brain_analysis",
            system_prompt=_BRAIN_SYSTEM_PROMPT,
            user_content=user_content,
            max_tokens=2048,
        )
        response = get_llm_router().generate(request)
        return response.raw_text or ""
    except Exception:
        return _fallback_analysis(data, analysis_type)


def _fallback_analysis(data: dict, analysis_type: str) -> str:
    """Rule-based fallback when LLM is unavailable."""
    if analysis_type == "entity":
        d = data.get("detail", {})
        trust = data.get("trust_score", 0)
        ctr = d.get("ctr", 0)
        cpc = d.get("cpc", 0)
        lines = [f"## Resumen\nEntidad con trust score {trust}/100."]
        if trust >= 70 and ctr > 1.5:
            lines.append("## Hallazgos Clave\n- Alto rendimiento con CTR fuerte")
            lines.append("## Recomendaciones Accionables\n- Considerar escalar presupuesto 20-30%")
            lines.append("## Siguiente Paso en el Flywheel\nOportunidades — escalar con nuevos angulos")
        elif trust < 40:
            lines.append(f"## Hallazgos Clave\n- Bajo rendimiento: CTR {ctr}%, CPC ${cpc:.2f}")
            lines.append("## Recomendaciones Accionables\n- Evaluar pausa o reestructuracion de audiencia")
            lines.append("## Siguiente Paso en el Flywheel\nBrand Profile — optimizar posicionamiento")
        else:
            lines.append(f"## Hallazgos Clave\n- Rendimiento moderado: CTR {ctr}%, CPC ${cpc:.2f}")
            lines.append("## Recomendaciones Accionables\n- A/B test con nuevos creativos")
            lines.append("## Siguiente Paso en el Flywheel\nCreativos — refrescar contenido publicitario")
        return "\n\n".join(lines)
    elif analysis_type == "feature":
        wr = round(data.get("win_rate", 0) * 100)
        lines = [f"## Resumen\nFeature con win rate de {wr}%."]
        if wr >= 60:
            lines.append("## Recomendaciones Accionables\n- Replicar esta estrategia en nuevas campanas")
        else:
            lines.append("## Recomendaciones Accionables\n- Probar nuevos angulos creativos para mejorar win rate")
        return "\n\n".join(lines)
    elif analysis_type == "outcome":
        label = data.get("outcome_label", "neutral")
        lines = [f"## Resumen\nResultado: {label}."]
        if label == "win":
            lines.append("## Lecciones Aprendidas\n- La estrategia actual funciona bien, documentar y replicar")
        elif label == "loss":
            lines.append("## Lecciones Aprendidas\n- Identificar factores de bajo rendimiento y ajustar")
        else:
            lines.append("## Lecciones Aprendidas\n- Rendimiento neutral, oportunidad de optimizacion")
        return "\n\n".join(lines)
    return ""


def _compute_flywheel_recommendations(
    entity_trust: List[EntityTrust],
    top_features: List[FeatureStats],
    summary: BrainSummary,
) -> list:
    """Rule-based flywheel routing recommendations."""
    recs = []

    # Analyze trust distribution
    low_trust = [e for e in entity_trust if e.trust_score < 40]
    high_trust = [e for e in entity_trust if e.trust_score > 70]

    # Low trust entities → Brand Profile
    if low_trust:
        avg_low_trust = round(sum(e.trust_score for e in low_trust) / len(low_trust), 1)
        recs.append({
            "module": "Brand Profile",
            "reason": f"{len(low_trust)} campanas con trust bajo (promedio {avg_low_trust}/100). Optimizar posicionamiento de marca para mejorar rendimiento base.",
            "priority": 1,
            "action_label": "Optimizar Perfil de Marca",
            "route_path": "/brand-profile",
        })

    # CTR declining → Creatives
    if summary.ctr_trend is not None and summary.ctr_trend < -10:
        recs.append({
            "module": "Creatives",
            "reason": f"CTR en declive {summary.ctr_trend}% vs periodo anterior. Refrescar creativos publicitarios con nuevos angulos y formatos.",
            "priority": 1,
            "action_label": "Crear Nuevos Creativos",
            "route_path": "/creatives",
        })

    # High CPC → Opportunities
    high_cpc_entities = [e for e in entity_trust if (e.detail or {}).get("cpc", 0) > 5 and (e.detail or {}).get("spend", 0) > 100]
    if high_cpc_entities:
        recs.append({
            "module": "Opportunities",
            "reason": f"{len(high_cpc_entities)} campanas con CPC alto. Analizar oportunidades para encontrar mejores angulos y reducir costos.",
            "priority": 2,
            "action_label": "Explorar Oportunidades",
            "route_path": "/opportunities",
        })

    # Low win rate features → Creatives
    low_wr_features = [f for f in top_features if f.win_rate < 0.3 and f.samples >= 2]
    if low_wr_features and not any(r["module"] == "Creatives" for r in recs):
        names = ", ".join(f.feature_key for f in low_wr_features[:3])
        recs.append({
            "module": "Creatives",
            "reason": f"Objetivos con bajo win rate: {names}. Probar nuevos creativos para estos objetivos.",
            "priority": 2,
            "action_label": "Refrescar Creativos",
            "route_path": "/creatives",
        })

    # High performers → Scale via Opportunities
    if high_trust:
        recs.append({
            "module": "Opportunities",
            "reason": f"{len(high_trust)} campanas de alto rendimiento (trust > 70). Escalar las mejores con nuevas oportunidades de mercado.",
            "priority": 3,
            "action_label": "Escalar Campanas",
            "route_path": "/opportunities",
        })

    # Always suggest saturation check if spending
    if summary.total_spend > 500:
        recs.append({
            "module": "Saturation",
            "reason": f"${summary.total_spend:,.0f} invertidos en el periodo. Verificar saturacion de audiencia y fatiga creativa.",
            "priority": 3,
            "action_label": "Analizar Saturacion",
            "route_path": "/saturation",
        })

    recs.sort(key=lambda r: r["priority"])
    return recs[:4]


# ── Deep Analysis Endpoints ──────────────────────────────────────────────────


@router.get("/brain/entity/{entity_id}/analysis")
def get_entity_analysis(
    entity_id: str,
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Deep LLM-powered analysis for a single entity."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)
    entities = _derive_entity_trust(db, org_uuid, since_dt, until_dt)

    target = None
    for e in entities:
        if e.entity_id == entity_id or (e.detail or {}).get("entity_meta_id") == entity_id:
            target = e
            break

    if not target:
        logger.warning("BRAIN_ENTITY_NOT_FOUND | entity_id={} | available={}", entity_id, [e.entity_id for e in entities[:5]])
        raise HTTPException(404, f"Entity '{entity_id}' not found")

    try:
        entity_dict = target.model_dump() if hasattr(target, 'model_dump') else target.dict()
        # Ensure datetime is serializable
        if entity_dict.get("last_seen_at") and not isinstance(entity_dict["last_seen_at"], str):
            entity_dict["last_seen_at"] = str(entity_dict["last_seen_at"])
        analysis_text = _generate_deep_analysis(entity_dict, "entity")

        d = target.detail or {}
        flywheel_next = None
        if target.trust_score < 40:
            flywheel_next = {"module": "Brand Profile", "reason": "Trust bajo, optimizar posicionamiento", "action_label": "Optimizar Marca", "route_path": "/brand-profile"}
        elif d.get("cpc", 0) > 5:
            flywheel_next = {"module": "Opportunities", "reason": "CPC alto, buscar mejores angulos", "action_label": "Explorar Oportunidades", "route_path": "/opportunities"}
        elif target.trust_score > 70:
            flywheel_next = {"module": "Opportunities", "reason": "Alto rendimiento, escalar", "action_label": "Escalar", "route_path": "/opportunities"}
        else:
            flywheel_next = {"module": "Creatives", "reason": "Rendimiento moderado, probar nuevos creativos", "action_label": "Crear Creativos", "route_path": "/creatives"}

        return {
            "entity": entity_dict,
            "analysis_text": analysis_text,
            "flywheel_next": flywheel_next,
        }
    except Exception as exc:
        logger.exception("BRAIN_ENTITY_ANALYSIS_ERROR | entity_id={} | error={}", entity_id, exc)
        # Return fallback response instead of 500
        return {
            "entity": {"entity_id": entity_id, "trust_score": target.trust_score, "detail": target.detail or {}},
            "analysis_text": _fallback_analysis({"trust_score": target.trust_score, "detail": target.detail or {}}, "entity"),
            "flywheel_next": {"module": "Opportunities", "reason": "Explorar opciones", "action_label": "Ver Oportunidades", "route_path": "/opportunities"},
        }


@router.get("/brain/feature/{feature_name}/analysis")
def get_feature_analysis(
    feature_name: str,
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Deep LLM-powered analysis for a single feature/objective."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)

    # Try derived features first
    features = _derive_top_features(db, org_uuid, since_dt, until_dt)
    target = None
    for f in features:
        if f.feature_key == feature_name:
            target = f
            break

    # Fallback: search FeatureMemory directly (learning tables may use different keys)
    if not target:
        mem = db.query(FeatureMemory).filter(
            FeatureMemory.org_id == org_uuid,
            FeatureMemory.feature_key == feature_name,
        ).first()
        if mem:
            target = FeatureStats(
                feature_type=mem.feature_type.value if hasattr(mem.feature_type, 'value') else str(mem.feature_type),
                feature_key=mem.feature_key,
                win_rate=mem.win_rate,
                samples=mem.samples,
                avg_delta=mem.avg_delta_json or {},
            )

    if not target:
        logger.warning("BRAIN_FEATURE_NOT_FOUND | feature_name={} | derived={} | memory_checked=True",
                        feature_name, [f.feature_key for f in features[:5]])
        raise HTTPException(404, f"Feature '{feature_name}' not found")

    try:
        feature_dict = target.model_dump() if hasattr(target, 'model_dump') else target.dict()
        analysis_text = _generate_deep_analysis(feature_dict, "feature")
        return {
            "feature": feature_dict,
            "analysis_text": analysis_text,
        }
    except Exception as exc:
        logger.exception("BRAIN_FEATURE_ANALYSIS_ERROR | feature_name={} | error={}", feature_name, exc)
        return {
            "feature": {"feature_key": feature_name, "win_rate": target.win_rate, "samples": target.samples},
            "analysis_text": _fallback_analysis({"win_rate": target.win_rate, "samples": target.samples}, "feature"),
        }


@router.get("/brain/outcome/{outcome_idx}/analysis")
def get_outcome_analysis(
    outcome_idx: int,
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Deep LLM-powered analysis for a single outcome."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)

    # Try learning tables first (same source as get_brain_stats)
    db_outcomes = db.query(DecisionOutcome).filter(
        DecisionOutcome.org_id == org_uuid,
        DecisionOutcome.outcome_label != OutcomeLabel.UNKNOWN,
    ).order_by(DecisionOutcome.created_at.desc()).limit(20).all()

    if db_outcomes:
        outcomes = [
            RecentOutcome(
                entity_type=o.entity_type,
                entity_id=o.entity_id,
                action_type=o.action_type.value if hasattr(o.action_type, 'value') else str(o.action_type),
                outcome_label=o.outcome_label.value if hasattr(o.outcome_label, 'value') else str(o.outcome_label),
                confidence=o.confidence,
                horizon_minutes=o.horizon_minutes,
                executed_at=o.executed_at,
            )
            for o in db_outcomes
        ]
    else:
        outcomes = _derive_recent_outcomes(db, org_uuid, since_dt, until_dt)

    if outcome_idx < 0 or outcome_idx >= len(outcomes):
        raise HTTPException(404, f"Outcome index {outcome_idx} not found")

    try:
        target = outcomes[outcome_idx]
        outcome_dict = target.model_dump() if hasattr(target, 'model_dump') else target.dict()
        if outcome_dict.get("executed_at") and not isinstance(outcome_dict["executed_at"], str):
            outcome_dict["executed_at"] = str(outcome_dict["executed_at"])
        analysis_text = _generate_deep_analysis(outcome_dict, "outcome")
    except Exception as exc:
        logger.exception("BRAIN_OUTCOME_ANALYSIS_ERROR | idx={} | error={}", outcome_idx, exc)
        outcome_dict = {"outcome_label": "neutral", "entity_id": "Unknown"}
        analysis_text = _fallback_analysis(outcome_dict, "outcome")

    return {
        "outcome": outcome_dict,
        "analysis_text": analysis_text,
    }


@router.get("/brain/flywheel")
def get_flywheel_recommendations(
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get flywheel routing recommendations based on current Brain intelligence."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)
    entity_trust = _derive_entity_trust(db, org_uuid, since_dt, until_dt)
    top_features = _derive_top_features(db, org_uuid, since_dt, until_dt)
    summary = _compute_summary(db, org_uuid, since_dt, until_dt, days)

    recs = _compute_flywheel_recommendations(entity_trust, top_features, summary)
    return {"recommendations": recs}


@router.get("/brain/entity/{entity_id}/pdf")
def export_entity_pdf(
    entity_id: str,
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export deep analysis PDF for a single entity."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)
    entities = _derive_entity_trust(db, org_uuid, since_dt, until_dt)

    target = None
    for e in entities:
        if e.entity_id == entity_id or (e.detail or {}).get("entity_meta_id") == entity_id:
            target = e
            break

    if not target:
        logger.warning("BRAIN_PDF_ENTITY_NOT_FOUND | entity_id={}", entity_id)
        raise HTTPException(404, f"Entity '{entity_id}' not found")

    try:
        entity_dict = target.model_dump() if hasattr(target, 'model_dump') else target.dict()
        # Convert datetime to string for PDF rendering
        if entity_dict.get("last_seen_at") and not isinstance(entity_dict["last_seen_at"], str):
            entity_dict["last_seen_at"] = str(entity_dict["last_seen_at"])
        analysis_text = _generate_deep_analysis(entity_dict, "entity")
        date_label = f"Last {days} days" if not since else f"{since} to {until or 'now'}"
        pdf_bytes = _build_entity_pdf(entity_dict, analysis_text, date_label)
    except Exception as exc:
        logger.exception("BRAIN_PDF_BUILD_ERROR | entity_id={} | error={}", entity_id, exc)
        raise HTTPException(500, f"PDF generation failed: {str(exc)}")

    safe_name = "".join(c if c.isalnum() or c in "-_ " else "" for c in entity_id[:20]).replace(" ", "_")
    filename = f"brain_entity_{safe_name}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Brain Export (PDF + XLSX) ────────────────────────────────────────────────


def _build_entity_pdf(entity_dict: dict, analysis_text: str, date_label: str) -> bytes:
    """Build a deep analysis PDF for a single entity."""
    from fpdf import FPDF
    from backend.src.utils.pdf_fonts import setup_pdf_fonts

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    _font = setup_pdf_fonts(pdf)

    pdf.add_page()

    # Title
    pdf.set_font(_font, "B", 18)
    pdf.cell(0, 12, "Entity Deep Analysis", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(_font, "", 10)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Period: {date_label}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Entity header
    det = entity_dict.get("detail", {})
    entity_name = entity_dict.get("entity_id", "Unknown")
    trust = entity_dict.get("trust_score", 0)

    pdf.set_font(_font, "B", 14)
    pdf.set_text_color(40, 40, 40)
    pdf.cell(0, 10, entity_name, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Trust breakdown box
    pdf.set_draw_color(212, 175, 55)
    pdf.set_fill_color(252, 249, 240)
    pdf.rect(10, pdf.get_y(), 190, 28, style="DF")
    y0 = pdf.get_y() + 3
    pdf.set_text_color(60, 60, 60)
    pdf.set_font(_font, "B", 10)
    pdf.set_xy(14, y0)
    pdf.cell(60, 7, f"Trust Score: {trust:.0f}/100")
    pdf.set_font(_font, "", 9)
    pdf.set_xy(14, y0 + 8)
    ctr_s = det.get("trust_ctr_score", 0)
    eff_s = det.get("trust_efficiency_score", 0)
    stab_s = det.get("trust_stability_score", 0)
    pdf.cell(60, 6, f"CTR Score: {ctr_s}/40")
    pdf.cell(60, 6, f"Efficiency: {eff_s}/30")
    pdf.cell(60, 6, f"Stability: {stab_s}/30")
    pdf.set_xy(14, y0 + 16)
    pdf.cell(60, 6, f"Objective: {det.get('objective', 'N/A')}")
    pdf.cell(60, 6, f"Status: {det.get('status', 'N/A')}")

    pdf.set_y(y0 + 28)
    pdf.ln(4)

    # Metrics table
    pdf.set_font(_font, "B", 11)
    pdf.set_text_color(40, 40, 40)
    pdf.cell(0, 8, "Performance Metrics", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font(_font, "B", 8)
    pdf.set_fill_color(240, 240, 235)
    cols = [32, 32, 32, 32, 32, 32]
    hdrs = ["Spend", "Clicks", "Impressions", "CTR", "CPC", "CPM"]
    for i, h in enumerate(hdrs):
        pdf.cell(cols[i], 6, h, border=1, fill=True)
    pdf.ln()
    pdf.set_font(_font, "", 8)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(cols[0], 5, f"${det.get('spend', 0):,.2f}", border=1)
    pdf.cell(cols[1], 5, f"{det.get('clicks', 0):,}", border=1)
    pdf.cell(cols[2], 5, f"{det.get('impressions', 0):,}", border=1)
    pdf.cell(cols[3], 5, f"{det.get('ctr', 0)}%", border=1)
    pdf.cell(cols[4], 5, f"${det.get('cpc', 0):.2f}", border=1)
    pdf.cell(cols[5], 5, f"${det.get('cpm', 0):.2f}", border=1)
    pdf.ln()
    pdf.ln(6)

    # LLM Analysis
    if analysis_text:
        pdf.set_font(_font, "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 8, "Deep Analysis", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Parse markdown sections
        for line in analysis_text.split("\n"):
            line = line.strip()
            if not line:
                pdf.ln(2)
                continue
            if line.startswith("## "):
                pdf.set_font(_font, "B", 10)
                pdf.set_text_color(40, 40, 40)
                pdf.cell(0, 7, line.replace("## ", "").replace("**", ""), new_x="LMARGIN", new_y="NEXT")
            elif line.startswith("- "):
                pdf.set_font(_font, "", 8)
                pdf.set_text_color(60, 60, 60)
                pdf.cell(5, 5, "")
                pdf.multi_cell(0, 4, line.replace("**", ""), new_x="LMARGIN", new_y="NEXT")
            else:
                pdf.set_font(_font, "", 8)
                pdf.set_text_color(60, 60, 60)
                pdf.multi_cell(0, 4, line.replace("**", ""), new_x="LMARGIN", new_y="NEXT")

    # Footer
    pdf.ln(6)
    pdf.set_font(_font, "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, "Generated by Meta Ops Agent - Brain Intelligence Engine", new_x="LMARGIN", new_y="NEXT")

    return pdf.output()


def _build_brain_pdf(summary_dict: dict, trust_list: list, features_list: list, suggestions_list: list, date_label: str) -> bytes:
    """Build Brain intelligence PDF report."""
    from fpdf import FPDF
    from backend.src.utils.pdf_fonts import setup_pdf_fonts

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    _font = setup_pdf_fonts(pdf)

    pdf.add_page()

    # Title
    pdf.set_font(_font, "B", 18)
    pdf.cell(0, 12, "Brain Intelligence Report", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(_font, "", 10)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Period: {date_label}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # KPI Summary Box
    pdf.set_draw_color(212, 175, 55)
    pdf.set_fill_color(252, 249, 240)
    pdf.rect(10, pdf.get_y(), 190, 24, style="DF")
    y0 = pdf.get_y() + 3
    pdf.set_text_color(60, 60, 60)
    pdf.set_font(_font, "B", 9)

    kpis = [
        [
            f"Campaigns: {summary_dict.get('total_campaigns', 0)}",
            f"Avg Trust: {summary_dict.get('avg_trust', 0)}",
            f"Winning: {summary_dict.get('win_count', 0)}/{summary_dict.get('total_campaigns', 0)}",
        ],
        [
            f"Spend: ${summary_dict.get('total_spend', 0):,.2f}",
            f"CTR: {summary_dict.get('avg_ctr', 0)}%",
            f"CPC: ${summary_dict.get('avg_cpc', 0):.2f}",
        ],
    ]
    for row_i, row in enumerate(kpis):
        pdf.set_xy(14, y0 + row_i * 9)
        for col_i, cell in enumerate(row):
            pdf.set_x(14 + col_i * 63)
            pdf.cell(63, 8, cell, new_x="RIGHT")

    pdf.set_y(y0 + 22)
    pdf.ln(4)

    # Entity Trust Table
    if trust_list:
        pdf.set_font(_font, "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 8, f"Entity Trust ({len(trust_list)})", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        pdf.set_font(_font, "B", 7)
        pdf.set_fill_color(240, 240, 235)
        cols = [50, 18, 22, 18, 18, 18, 18, 18, 12]
        hdrs = ["Entity", "Trust", "Objective", "Spend", "CTR", "CPC", "CPM", "Clicks", "Days"]
        for i, h in enumerate(hdrs):
            pdf.cell(cols[i], 6, h, border=1, fill=True)
        pdf.ln()

        pdf.set_font(_font, "", 7)
        pdf.set_text_color(60, 60, 60)
        for e in trust_list[:30]:
            d = e if isinstance(e, dict) else (e.detail if hasattr(e, 'detail') else {})
            eid = e.get("entity_id", "") if isinstance(e, dict) else getattr(e, "entity_id", "")
            trust = e.get("trust_score", 0) if isinstance(e, dict) else getattr(e, "trust_score", 0)
            det = d.get("detail", d) if isinstance(d, dict) else {}
            if isinstance(e, dict) and "detail" in e:
                det = e["detail"]
            elif hasattr(e, "detail"):
                det = e.detail or {}

            pdf.cell(cols[0], 5, str(eid)[:26], border=1)
            pdf.cell(cols[1], 5, f"{trust:.0f}", border=1)
            pdf.cell(cols[2], 5, str(det.get("objective", ""))[:12], border=1)
            pdf.cell(cols[3], 5, f"${det.get('spend', 0):,.0f}", border=1)
            pdf.cell(cols[4], 5, f"{det.get('ctr', 0)}%", border=1)
            pdf.cell(cols[5], 5, f"${det.get('cpc', 0):.2f}", border=1)
            pdf.cell(cols[6], 5, f"${det.get('cpm', 0):.2f}", border=1)
            pdf.cell(cols[7], 5, f"{det.get('clicks', 0):,}", border=1)
            pdf.cell(cols[8], 5, str(det.get("days_active", "")), border=1)
            pdf.ln()
        pdf.ln(4)

    # Top Features
    if features_list:
        pdf.set_font(_font, "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 8, f"Top Features ({len(features_list)})", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        for f in features_list[:15]:
            fdata = f if isinstance(f, dict) else {"feature_key": getattr(f, "feature_key", ""), "feature_type": getattr(f, "feature_type", ""), "win_rate": getattr(f, "win_rate", 0), "samples": getattr(f, "samples", 0)}
            pct = round(fdata.get("win_rate", 0) * 100)
            pdf.set_font(_font, "B", 8)
            pdf.set_text_color(60, 60, 60)
            pdf.cell(60, 5, f"{fdata.get('feature_key', '')} ({fdata.get('feature_type', '')})")
            pdf.set_font(_font, "", 8)
            pdf.cell(30, 5, f"Win: {pct}%")
            pdf.cell(30, 5, f"Samples: {fdata.get('samples', 0)}")
            pdf.ln()
        pdf.ln(4)

    # Suggestions
    if suggestions_list:
        pdf.set_font(_font, "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 8, f"Suggested Actions ({len(suggestions_list)})", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        type_prefix = {"scale": "[SCALE]", "pause": "[PAUSE]", "test": "[TEST]", "refresh": "[REFRESH]", "optimize": "[OPT]"}
        for s in suggestions_list:
            sdata = s if isinstance(s, dict) else s.__dict__
            stype = sdata.get("type", "")
            pdf.set_font(_font, "B", 9)
            pdf.set_text_color(60, 60, 60)
            pdf.cell(0, 6, f"{type_prefix.get(stype, '')} {sdata.get('title', '')}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(_font, "", 8)
            pdf.set_text_color(100, 100, 100)
            pdf.multi_cell(0, 4, sdata.get("description", ""), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

    # Footer
    pdf.ln(6)
    pdf.set_font(_font, "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, "Generated by Meta Ops Agent - Brain Intelligence Engine", new_x="LMARGIN", new_y="NEXT")

    return pdf.output()


def _build_brain_xlsx(summary_dict: dict, trust_list: list, features_list: list, suggestions_list: list, date_label: str) -> bytes:
    """Build Brain XLSX with four sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    hfont = Font(bold=True, size=11)
    hfill = PatternFill(start_color="F0F0EB", end_color="F0F0EB", fill_type="solid")

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Brain Intelligence Report", "", date_label])
    ws1.append([])
    ws1.append(["Metric", "Value"])
    for cell in ws1[3]:
        cell.font = hfont
        cell.fill = hfill
    for k in ["total_campaigns", "avg_trust", "win_count", "loss_count", "total_spend", "total_clicks", "total_impressions", "avg_ctr", "avg_cpc", "avg_cpm"]:
        ws1.append([k.replace("_", " ").title(), summary_dict.get(k, 0)])
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18

    # Sheet 2: Entity Trust
    ws2 = wb.create_sheet("Entity Trust")
    ws2.append(["Entity", "Trust Score", "Last Outcome", "Objective", "Status", "Spend", "CTR", "CPC", "CPM", "Clicks", "Impressions", "Days Active"])
    for cell in ws2[1]:
        cell.font = hfont
        cell.fill = hfill
    for e in trust_list:
        ed = e if isinstance(e, dict) else {"entity_id": getattr(e, "entity_id", ""), "trust_score": getattr(e, "trust_score", 0), "last_outcome": getattr(e, "last_outcome", ""), "detail": getattr(e, "detail", {})}
        det = ed.get("detail", {})
        ws2.append([
            ed.get("entity_id", ""), ed.get("trust_score", 0), ed.get("last_outcome", ""),
            det.get("objective", ""), det.get("status", ""), det.get("spend", 0),
            det.get("ctr", 0), det.get("cpc", 0), det.get("cpm", 0),
            det.get("clicks", 0), det.get("impressions", 0), det.get("days_active", 0),
        ])
    ws2.column_dimensions["A"].width = 30

    # Sheet 3: Features
    ws3 = wb.create_sheet("Features")
    ws3.append(["Feature Type", "Feature Key", "Win Rate", "Samples", "Spend", "Clicks", "Avg CTR", "Avg CPC"])
    for cell in ws3[1]:
        cell.font = hfont
        cell.fill = hfill
    for f in features_list:
        fd = f if isinstance(f, dict) else {"feature_type": getattr(f, "feature_type", ""), "feature_key": getattr(f, "feature_key", ""), "win_rate": getattr(f, "win_rate", 0), "samples": getattr(f, "samples", 0), "avg_delta": getattr(f, "avg_delta", {})}
        ad = fd.get("avg_delta", {})
        ws3.append([
            fd.get("feature_type", ""), fd.get("feature_key", ""), fd.get("win_rate", 0), fd.get("samples", 0),
            ad.get("spend", 0), ad.get("clicks", 0), ad.get("avg_ctr", 0), ad.get("avg_cpc", 0),
        ])
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 22

    # Sheet 4: Suggestions
    ws4 = wb.create_sheet("Suggestions")
    ws4.append(["Type", "Entity", "Title", "Description"])
    for cell in ws4[1]:
        cell.font = hfont
        cell.fill = hfill
    for s in suggestions_list:
        sd = s if isinstance(s, dict) else s.__dict__
        ws4.append([sd.get("type", ""), sd.get("entity_id", ""), sd.get("title", ""), sd.get("description", "")])
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 25
    ws4.column_dimensions["C"].width = 40
    ws4.column_dimensions["D"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/brain/export/pdf")
def export_brain_pdf(
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export Brain intelligence data as PDF."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)

    try:
        summary = _compute_summary(db, org_uuid, since_dt, until_dt, days)
        entity_trust = _derive_entity_trust(db, org_uuid, since_dt, until_dt)
        top_features = _derive_top_features(db, org_uuid, since_dt, until_dt)

        # Fill summary from trust data
        summary.total_campaigns = len(entity_trust)
        summary.win_count = sum(1 for e in entity_trust if e.last_outcome == "win")
        summary.loss_count = sum(1 for e in entity_trust if e.last_outcome == "loss")
        summary.avg_trust = round(sum(e.trust_score for e in entity_trust) / len(entity_trust), 1) if entity_trust else 0

        # Get suggestions
        suggestions_resp = get_brain_suggestions(days=days, since=since, until=until, db=db, user=user)
        suggestions_list = suggestions_resp.get("suggestions", [])

        date_label = f"Last {days} days" if not since else f"{since} to {until or 'now'}"

        # Convert pydantic models to dicts for the builder
        summary_dict = summary.dict() if hasattr(summary, 'dict') else summary.__dict__
        trust_dicts = [e.dict() if hasattr(e, 'dict') else e.__dict__ for e in entity_trust]
        features_dicts = [f.dict() if hasattr(f, 'dict') else f.__dict__ for f in top_features]

        pdf_bytes = _build_brain_pdf(summary_dict, trust_dicts, features_dicts, suggestions_list, date_label)

        filename = f"brain_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate Brain PDF: {str(e)}")


@router.get("/brain/export/xlsx")
def export_brain_xlsx(
    days: int = Query(90, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export Brain intelligence data as XLSX."""
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since_dt, until_dt = _parse_dates(days, since, until)

    try:
        summary = _compute_summary(db, org_uuid, since_dt, until_dt, days)
        entity_trust = _derive_entity_trust(db, org_uuid, since_dt, until_dt)
        top_features = _derive_top_features(db, org_uuid, since_dt, until_dt)

        summary.total_campaigns = len(entity_trust)
        summary.win_count = sum(1 for e in entity_trust if e.last_outcome == "win")
        summary.loss_count = sum(1 for e in entity_trust if e.last_outcome == "loss")
        summary.avg_trust = round(sum(e.trust_score for e in entity_trust) / len(entity_trust), 1) if entity_trust else 0

        suggestions_resp = get_brain_suggestions(days=days, since=since, until=until, db=db, user=user)
        suggestions_list = suggestions_resp.get("suggestions", [])

        date_label = f"Last {days} days" if not since else f"{since} to {until or 'now'}"

        summary_dict = summary.dict() if hasattr(summary, 'dict') else summary.__dict__
        trust_dicts = [e.dict() if hasattr(e, 'dict') else e.__dict__ for e in entity_trust]
        features_dicts = [f.dict() if hasattr(f, 'dict') else f.__dict__ for f in top_features]

        xlsx_bytes = _build_brain_xlsx(summary_dict, trust_dicts, features_dicts, suggestions_list, date_label)

        filename = f"brain_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate Brain XLSX: {str(e)}")


# ── Creative Outcome Capture ──────────────────────────────────────────────────


@router.post("/brain/capture-creative-outcomes")
def capture_creative_outcomes(
    days: int = Query(14, le=90),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """
    Evaluate creative performance from Meta insights and feed results back into Brain memory.
    Compares recent performance vs baseline for each creative (ad-level), labels WIN/LOSS/NEUTRAL,
    and updates FeatureMemory with creative tags.
    """
    from backend.src.database.models import Creative, FeatureMemory, FeatureType
    org_id = user.get("org_id", "")
    try:
        org_uuid = UUID(org_id)
    except (ValueError, AttributeError):
        raise HTTPException(400, "Invalid organization context")

    since = datetime.utcnow() - timedelta(days=days)
    half = datetime.utcnow() - timedelta(days=days // 2)

    # Get all creatives that have a linked Meta ad
    creatives = db.query(Creative).filter(
        Creative.meta_ad_id.isnot(None),
    ).all()

    if not creatives:
        return {"status": "no_creatives", "processed": 0, "wins": 0, "losses": 0}

    wins = 0
    losses = 0
    neutrals = 0
    processed = 0

    for creative in creatives:
        # Get insights for this ad in the period
        insights = db.query(MetaInsightsDaily).filter(
            MetaInsightsDaily.org_id == org_uuid,
            MetaInsightsDaily.entity_meta_id == creative.meta_ad_id,
            MetaInsightsDaily.level == InsightLevel.AD,
            MetaInsightsDaily.date_start >= since,
        ).order_by(MetaInsightsDaily.date_start).all()

        if len(insights) < 4:
            continue

        # Split into baseline (first half) vs recent (second half)
        baseline = [r for r in insights if r.date_start < half]
        recent = [r for r in insights if r.date_start >= half]

        if not baseline or not recent:
            continue

        # Compute averages
        baseline_ctr = sum(r.ctr or 0 for r in baseline) / len(baseline)
        recent_ctr = sum(r.ctr or 0 for r in recent) / len(recent)
        baseline_cpc = sum((r.spend or 0) / max(r.clicks or 1, 1) for r in baseline) / len(baseline)
        recent_cpc = sum((r.spend or 0) / max(r.clicks or 1, 1) for r in recent) / len(recent)

        # Label
        ctr_delta = recent_ctr - baseline_ctr
        cpc_delta = recent_cpc - baseline_cpc

        if ctr_delta > 0.1 and cpc_delta <= 0:
            label = "win"
            wins += 1
        elif ctr_delta < -0.2 or cpc_delta > baseline_cpc * 0.3:
            label = "loss"
            losses += 1
        else:
            label = "neutral"
            neutrals += 1

        processed += 1

        # Update FeatureMemory for creative tags
        tags = creative.tags or []
        for tag_obj in tags:
            tag_key = tag_obj.get("l1", "") if isinstance(tag_obj, dict) else str(tag_obj)
            if not tag_key:
                continue

            feature = db.query(FeatureMemory).filter(
                FeatureMemory.org_id == org_uuid,
                FeatureMemory.feature_type == FeatureType.TAG,
                FeatureMemory.feature_key == tag_key,
            ).first()

            if not feature:
                feature = FeatureMemory(
                    org_id=org_uuid,
                    feature_type=FeatureType.TAG,
                    feature_key=tag_key,
                    win_rate=0.0,
                    samples=0,
                    avg_delta_json={},
                )
                db.add(feature)
                db.flush()

            # Update win rate with rolling average
            old_count = feature.samples or 0
            new_count = old_count + 1
            is_win = 1.0 if label == "win" else 0.0
            feature.win_rate = round(
                (feature.win_rate * old_count + is_win) / new_count, 4
            )
            feature.samples = new_count
            feature.avg_delta_json = {
                **(feature.avg_delta_json or {}),
                "last_ctr_delta": round(ctr_delta, 4),
                "last_cpc_delta": round(cpc_delta, 4),
                "last_label": label,
                "last_evaluated": datetime.utcnow().isoformat(),
            }

    db.commit()

    logger.info(
        "CREATIVE_OUTCOME_CAPTURE | org={} | processed={} | wins={} | losses={} | neutrals={}",
        org_id, processed, wins, losses, neutrals,
    )

    return {
        "status": "success",
        "processed": processed,
        "wins": wins,
        "losses": losses,
        "neutrals": neutrals,
        "total_creatives": len(creatives),
    }
