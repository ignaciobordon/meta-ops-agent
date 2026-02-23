"""Analytics API: metrics-over-time, insights, enriched summary & campaigns."""
from typing import Optional
from datetime import datetime
from uuid import UUID
import io

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.src.database.session import get_db
from backend.src.middleware.auth import get_current_user
from backend.src.services import analytics_service, benchmark_service

router = APIRouter()


@router.get("/summary")
def get_performance_summary(
    days: int = Query(7, le=730),
    since: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    until: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    ad_account_id: Optional[str] = Query(None, description="Filter by ad account UUID"),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get aggregated performance summary with period-over-period trends."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    acc_id = UUID(ad_account_id) if ad_account_id else None
    return analytics_service.get_performance_summary(db, UUID(org_id), days, since, until, acc_id)


@router.get("/metrics-over-time")
def get_metrics_over_time(
    days: int = Query(30, le=730),
    since: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    until: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    ad_account_id: Optional[str] = Query(None, description="Filter by ad account UUID"),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get bucketed metrics over time (daily/weekly/monthly based on span)."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    acc_id = UUID(ad_account_id) if ad_account_id else None
    return analytics_service.get_metrics_over_time(db, UUID(org_id), days, since, until, acc_id)


@router.get("/insights")
def get_insights(
    days: int = Query(30, le=730),
    since: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    until: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    ad_account_id: Optional[str] = Query(None, description="Filter by ad account UUID"),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get rule-based analytical insights."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    acc_id = UUID(ad_account_id) if ad_account_id else None
    return {"insights": analytics_service.generate_insights(db, UUID(org_id), days, since, until, acc_id)}


@router.get("/spend-over-time")
def get_spend_over_time(
    days: int = Query(30, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    ad_account_id: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get daily spend data for charting (legacy)."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    acc_id = UUID(ad_account_id) if ad_account_id else None
    return analytics_service.get_spend_over_time(db, UUID(org_id), days, since, until, acc_id)


@router.get("/top-campaigns")
def get_top_campaigns(
    days: int = Query(7, le=730),
    limit: int = Query(20, le=50),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    ad_account_id: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get top campaigns by spend with objective, status, and enriched metrics."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    acc_id = UUID(ad_account_id) if ad_account_id else None
    return analytics_service.get_top_campaigns(db, UUID(org_id), days, limit, since, until, acc_id)


@router.get("/daily")
def get_daily_breakdown(
    days: int = Query(30, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    ad_account_id: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get daily breakdown of all key metrics."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    acc_id = UUID(ad_account_id) if ad_account_id else None
    return analytics_service.get_daily_breakdown(db, UUID(org_id), days, since, until, acc_id)


@router.get("/benchmarks")
def get_benchmarks(
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get org vs own 30-day baseline benchmarks."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    return benchmark_service.get_benchmarks(db, UUID(org_id))


# ── Analytics Export (PDF + XLSX) ─────────────────────────────────────────────


def _build_analytics_pdf(summary: dict, campaigns: list, insights: list, date_label: str) -> bytes:
    """Build analytics PDF report from summary + campaigns + insights."""
    from fpdf import FPDF
    from backend.src.utils.pdf_fonts import setup_pdf_fonts

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    _font = setup_pdf_fonts(pdf)

    pdf.add_page()

    # Title
    pdf.set_font(_font, "B", 18)
    pdf.cell(0, 12, "Analytics Report", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(_font, "", 10)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Period: {date_label}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # KPI Summary Box
    pdf.set_draw_color(212, 175, 55)
    pdf.set_fill_color(252, 249, 240)
    pdf.rect(10, pdf.get_y(), 190, 32, style="DF")
    y0 = pdf.get_y() + 3
    pdf.set_text_color(60, 60, 60)
    pdf.set_font(_font, "B", 9)

    kpis = [
        [
            f"Spend: ${summary.get('total_spend', 0):,.2f}",
            f"Clicks: {summary.get('total_clicks', 0):,}",
            f"Impressions: {summary.get('total_impressions', 0):,}",
        ],
        [
            f"CTR: {summary.get('avg_ctr', 0)}%",
            f"CPC: ${summary.get('avg_cpc', 0):.2f}",
            f"CPM: ${summary.get('avg_cpm', 0):.2f}",
        ],
        [
            f"Conversions: {summary.get('total_conversions', 0):,}",
            f"ROAS: {summary.get('avg_roas', 0)}x",
            f"Active Campaigns: {summary.get('active_campaigns', 0)}",
        ],
    ]
    for row_i, row in enumerate(kpis):
        pdf.set_xy(14, y0 + row_i * 9)
        for col_i, cell in enumerate(row):
            pdf.set_x(14 + col_i * 63)
            pdf.cell(63, 8, cell, new_x="RIGHT")

    pdf.set_y(y0 + 30)
    pdf.ln(4)

    # Trends
    trends = []
    for key in ["spend_trend", "ctr_trend", "cpc_trend", "roas_trend", "clicks_trend", "impressions_trend", "conversions_trend"]:
        val = summary.get(key)
        if val is not None:
            label = key.replace("_trend", "").replace("_", " ").upper()
            trends.append(f"{label}: {'+' if val > 0 else ''}{val}%")

    if trends:
        pdf.set_font(_font, "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 8, "Period Trends", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font(_font, "", 9)
        pdf.set_text_color(80, 80, 80)
        pdf.multi_cell(0, 5, "  |  ".join(trends), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

    # Top Campaigns Table
    if campaigns:
        pdf.set_font(_font, "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 8, f"Top Campaigns ({len(campaigns)})", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Header
        pdf.set_font(_font, "B", 7)
        pdf.set_fill_color(240, 240, 235)
        col_widths = [55, 22, 18, 22, 15, 15, 15, 15, 15]
        headers = ["Campaign", "Objective", "Status", "Spend", "CTR", "CPC", "ROAS", "Clicks", "Impr."]
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 6, h, border=1, fill=True)
        pdf.ln()

        # Rows
        pdf.set_font(_font, "", 7)
        pdf.set_text_color(60, 60, 60)
        for c in campaigns[:30]:
            name = (c.get("name", "") or "")[:28]
            pdf.cell(col_widths[0], 5, name, border=1)
            pdf.cell(col_widths[1], 5, str(c.get("objective", ""))[:12], border=1)
            pdf.cell(col_widths[2], 5, str(c.get("status", ""))[:10], border=1)
            pdf.cell(col_widths[3], 5, f"${c.get('spend', 0):,.0f}", border=1)
            pdf.cell(col_widths[4], 5, f"{c.get('ctr', 0)}%", border=1)
            pdf.cell(col_widths[5], 5, f"${c.get('cpc', 0):.2f}", border=1)
            pdf.cell(col_widths[6], 5, f"{c.get('roas', 0)}x", border=1)
            pdf.cell(col_widths[7], 5, f"{c.get('clicks', 0):,}", border=1)
            pdf.cell(col_widths[8], 5, f"{c.get('impressions', 0):,}", border=1)
            pdf.ln()
        pdf.ln(4)

    # Insights
    if insights:
        pdf.set_font(_font, "B", 11)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 8, f"Insights ({len(insights)})", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        for ins in insights:
            itype = ins.get("type", "info")
            prefix = "[+]" if itype == "positive" else "[!]" if itype == "warning" else "[i]"
            pdf.set_font(_font, "B", 9)
            pdf.set_text_color(60, 60, 60)
            pdf.cell(0, 6, f"{prefix} {ins.get('title', '')}  ({ins.get('metric_value', '')})", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(_font, "", 8)
            pdf.set_text_color(100, 100, 100)
            pdf.multi_cell(0, 4, ins.get("description", ""), new_x="LMARGIN", new_y="NEXT")
            pdf.ln(2)

    # Footer
    pdf.ln(6)
    pdf.set_font(_font, "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, "Generated by Meta Ops Agent - Analytics Engine", new_x="LMARGIN", new_y="NEXT")

    return pdf.output()


def _build_analytics_xlsx(summary: dict, campaigns: list, insights: list, date_label: str) -> bytes:
    """Build analytics XLSX with three sheets: Overview, Campaigns, Insights."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="F0F0EB", end_color="F0F0EB", fill_type="solid")

    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.append(["Analytics Report", "", date_label])
    ws1.append([])
    ws1.append(["Metric", "Value", "Trend %"])
    for cell in ws1[3]:
        cell.font = header_font
        cell.fill = header_fill

    metrics = [
        ("Total Spend", summary.get("total_spend", 0), summary.get("spend_trend")),
        ("Total Clicks", summary.get("total_clicks", 0), summary.get("clicks_trend")),
        ("Total Impressions", summary.get("total_impressions", 0), summary.get("impressions_trend")),
        ("Total Conversions", summary.get("total_conversions", 0), summary.get("conversions_trend")),
        ("Avg CTR (%)", summary.get("avg_ctr", 0), summary.get("ctr_trend")),
        ("Avg CPC ($)", summary.get("avg_cpc", 0), summary.get("cpc_trend")),
        ("Avg CPM ($)", summary.get("avg_cpm", 0), None),
        ("Avg ROAS", summary.get("avg_roas", 0), summary.get("roas_trend")),
        ("Active Campaigns", summary.get("active_campaigns", 0), None),
    ]
    for label, val, trend in metrics:
        ws1.append([label, val, trend])

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 12

    # Sheet 2: Campaigns
    ws2 = wb.create_sheet("Campaigns")
    camp_headers = ["Campaign", "Objective", "Status", "Spend", "Clicks", "Impressions", "Conversions", "CTR", "CPC", "CPM", "ROAS", "Frequency"]
    ws2.append(camp_headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    for c in campaigns:
        ws2.append([
            c.get("name", ""), c.get("objective", ""), c.get("status", ""),
            c.get("spend", 0), c.get("clicks", 0), c.get("impressions", 0),
            c.get("conversions", 0), c.get("ctr", 0), c.get("cpc", 0),
            c.get("cpm", 0), c.get("roas", 0), c.get("frequency", 0),
        ])

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 20

    # Sheet 3: Insights
    ws3 = wb.create_sheet("Insights")
    ws3.append(["Type", "Title", "Description", "Metric Value"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill

    for ins in insights:
        ws3.append([ins.get("type", ""), ins.get("title", ""), ins.get("description", ""), ins.get("metric_value", "")])

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 60
    ws3.column_dimensions["D"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/export/pdf")
def export_analytics_pdf(
    days: int = Query(30, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    ad_account_id: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export analytics dashboard data as a PDF report."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    org_uuid = UUID(org_id)
    acc_id = UUID(ad_account_id) if ad_account_id else None

    try:
        summary = analytics_service.get_performance_summary(db, org_uuid, days, since, until, acc_id)
        campaigns = analytics_service.get_top_campaigns(db, org_uuid, days, 30, since, until, acc_id)
        insights_resp = analytics_service.generate_insights(db, org_uuid, days, since, until, acc_id)

        date_label = f"Last {days} days" if not since else f"{since} to {until or 'now'}"

        summary_dict = summary if isinstance(summary, dict) else (summary.__dict__ if hasattr(summary, '__dict__') else {})
        campaigns_list = campaigns if isinstance(campaigns, list) else []
        insights_list = insights_resp if isinstance(insights_resp, list) else []

        pdf_bytes = _build_analytics_pdf(summary_dict, campaigns_list, insights_list, date_label)

        filename = f"analytics_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate analytics PDF: {str(e)}")


@router.get("/export/xlsx")
def export_analytics_xlsx(
    days: int = Query(30, le=730),
    since: Optional[str] = Query(None),
    until: Optional[str] = Query(None),
    ad_account_id: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export analytics dashboard data as an XLSX workbook."""
    org_id = user.get("org_id")
    if not org_id:
        raise HTTPException(400, "No organization")

    org_uuid = UUID(org_id)
    acc_id = UUID(ad_account_id) if ad_account_id else None

    try:
        summary = analytics_service.get_performance_summary(db, org_uuid, days, since, until, acc_id)
        campaigns = analytics_service.get_top_campaigns(db, org_uuid, days, 30, since, until, acc_id)
        insights_resp = analytics_service.generate_insights(db, org_uuid, days, since, until, acc_id)

        date_label = f"Last {days} days" if not since else f"{since} to {until or 'now'}"

        summary_dict = summary if isinstance(summary, dict) else (summary.__dict__ if hasattr(summary, '__dict__') else {})
        campaigns_list = campaigns if isinstance(campaigns, list) else []
        insights_list = insights_resp if isinstance(insights_resp, list) else []

        xlsx_bytes = _build_analytics_xlsx(summary_dict, campaigns_list, insights_list, date_label)

        filename = f"analytics_report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate analytics XLSX: {str(e)}")
