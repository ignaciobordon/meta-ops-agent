"""Content Studio API — Production Pack generation, management, export."""
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
from pydantic import BaseModel, Field
from uuid import UUID
from datetime import datetime
import io

from backend.src.database.session import get_db
from backend.src.database.models import (
    ContentPack, ContentVariant, ContentChannelLock, ContentExport, ContentPackStatus,
)
from backend.src.middleware.auth import get_current_user, require_operator_or_admin
from backend.src.content.channel_specs import CHANNEL_SPECS, get_all_channels
from src.utils.logging_config import set_trace_id, logger
from uuid import uuid4

router = APIRouter(tags=["content-studio"])


# ── Request/Response Models ──────────────────────────────────────────────────

class ChannelSelection(BaseModel):
    channel: str
    format: str = ""

class CreatePackRequest(BaseModel):
    creative_id: str
    channels: List[ChannelSelection]
    goal: str = "awareness"
    language: str = "es-AR"
    target_audience: str = ""
    tone_tags: List[str] = []
    curator_prompt: str = ""
    framework_preference: str = ""
    hook_style_preference: str = ""
    compliance_mode: str = "safe"
    brand_voice_mode: str = "strict"

class AsyncJobResponse(BaseModel):
    pack_id: str
    job_id: str
    status: str = "queued"

class VariantResponse(BaseModel):
    id: str
    channel: str
    format: str = ""
    variant_index: int
    output_json: Dict[str, Any] = {}
    score: float = 0.0
    score_breakdown_json: Dict[str, Any] = {}
    rationale_text: str = ""

class PackResponse(BaseModel):
    id: str
    creative_id: str = ""
    status: str
    goal: str = ""
    language: str = "es-AR"
    channels_json: List[Dict[str, Any]] = []
    input_json: Dict[str, Any] = {}
    last_error_code: str = ""
    last_error_message: str = ""
    created_at: str
    variants_count: int = 0

class LockRequest(BaseModel):
    channel: str
    variant_id: str

class RegenerateRequest(BaseModel):
    channels: List[str] = []  # empty = all unlocked channels

class ChannelSpecResponse(BaseModel):
    key: str
    display_name: str
    platform: str
    default_format: str
    variants_count: int
    format_options: List[str]
    required_fields: List[str]


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/channels", response_model=List[ChannelSpecResponse])
def list_channels(user: dict = Depends(get_current_user)):
    """List all available channels and their specs."""
    return [
        ChannelSpecResponse(
            key=s.key,
            display_name=s.display_name,
            platform=s.platform,
            default_format=s.default_format,
            variants_count=s.variants_count,
            format_options=s.format_options,
            required_fields=s.required_fields,
        )
        for s in get_all_channels()
    ]


@router.post("/packs", status_code=202, response_model=AsyncJobResponse)
def create_pack(
    request: CreatePackRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Create a content pack and enqueue generation job."""
    org_id = user.get("org_id", "")
    if not org_id:
        raise HTTPException(400, "No organization context")

    if not request.channels:
        raise HTTPException(400, "At least one channel is required")

    # Validate channels
    for ch in request.channels:
        if ch.channel not in CHANNEL_SPECS:
            raise HTTPException(400, f"Unknown channel: {ch.channel}")

    from backend.src.services.content_creator_service import build_pack_from_creative
    from backend.src.jobs.queue import enqueue

    settings = {
        "goal": request.goal,
        "language": request.language,
        "target_audience": request.target_audience,
        "tone_tags": request.tone_tags,
        "curator_prompt": request.curator_prompt,
        "framework_preference": request.framework_preference,
        "hook_style_preference": request.hook_style_preference,
        "compliance_mode": request.compliance_mode,
        "brand_voice_mode": request.brand_voice_mode,
    }

    channels = [{"channel": ch.channel, "format": ch.format} for ch in request.channels]

    pack = build_pack_from_creative(
        db=db,
        org_id=org_id,
        creative_id=request.creative_id,
        channels=channels,
        settings=settings,
    )

    job_id = enqueue(
        task_name="content_studio_generate",
        payload={"pack_id": str(pack.id)},
        org_id=org_id,
        db=db,
    )

    pack.job_run_id = UUID(job_id)
    db.commit()

    logger.info(
        "CONTENT_STUDIO_PACK_CREATED | pack={} | job={} | channels={}",
        pack.id, job_id, len(channels),
    )

    return AsyncJobResponse(pack_id=str(pack.id), job_id=job_id, status="queued")


@router.get("/packs", response_model=List[PackResponse])
def list_packs(
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """List all content packs for the current organization."""
    org_id = user.get("org_id", "")
    if not org_id:
        raise HTTPException(400, "No organization context")

    packs = db.query(ContentPack).filter(
        ContentPack.org_id == UUID(org_id)
    ).order_by(ContentPack.created_at.desc()).limit(limit).all()

    results = []
    for pack in packs:
        variants_count = db.query(ContentVariant).filter(
            ContentVariant.content_pack_id == pack.id
        ).count()

        results.append(PackResponse(
            id=str(pack.id),
            creative_id=str(pack.creative_id) if pack.creative_id else "",
            status=pack.status.value if pack.status else "queued",
            goal=pack.goal or "",
            language=pack.language or "es-AR",
            channels_json=pack.channels_json or [],
            input_json=pack.input_json or {},
            last_error_code=pack.last_error_code or "",
            last_error_message=pack.last_error_message or "",
            created_at=pack.created_at.isoformat() if pack.created_at else "",
            variants_count=variants_count,
        ))

    return results


@router.get("/packs/{pack_id}", response_model=PackResponse)
def get_pack(
    pack_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get a content pack by ID."""
    org_id = user.get("org_id", "")
    pack = db.query(ContentPack).filter(
        ContentPack.id == UUID(pack_id),
        ContentPack.org_id == UUID(org_id),
    ).first()

    if not pack:
        raise HTTPException(404, "Content pack not found")

    variants_count = db.query(ContentVariant).filter(
        ContentVariant.content_pack_id == pack.id
    ).count()

    return PackResponse(
        id=str(pack.id),
        creative_id=str(pack.creative_id) if pack.creative_id else "",
        status=pack.status.value if pack.status else "queued",
        goal=pack.goal or "",
        language=pack.language or "es-AR",
        channels_json=pack.channels_json or [],
        input_json=pack.input_json or {},
        last_error_code=pack.last_error_code or "",
        last_error_message=pack.last_error_message or "",
        created_at=pack.created_at.isoformat() if pack.created_at else "",
        variants_count=variants_count,
    )


@router.get("/packs/{pack_id}/variants", response_model=List[VariantResponse])
def get_variants(
    pack_id: str,
    channel: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get variants for a pack, optionally filtered by channel."""
    org_id = user.get("org_id", "")

    # Verify pack belongs to org
    pack = db.query(ContentPack).filter(
        ContentPack.id == UUID(pack_id),
        ContentPack.org_id == UUID(org_id),
    ).first()
    if not pack:
        raise HTTPException(404, "Content pack not found")

    query = db.query(ContentVariant).filter(ContentVariant.content_pack_id == pack.id)
    if channel:
        query = query.filter(ContentVariant.channel == channel)

    variants = query.order_by(ContentVariant.channel, ContentVariant.variant_index).all()

    return [
        VariantResponse(
            id=str(v.id),
            channel=v.channel,
            format=v.format or "",
            variant_index=v.variant_index,
            output_json=v.output_json or {},
            score=v.score or 0.0,
            score_breakdown_json=v.score_breakdown_json or {},
            rationale_text=v.rationale_text or "",
        )
        for v in variants
    ]


@router.post("/packs/{pack_id}/lock")
def lock_variant(
    pack_id: str,
    request: LockRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Lock a variant as the final selection for a channel."""
    org_id = user.get("org_id", "")
    user_id = user.get("user_id", "")

    pack = db.query(ContentPack).filter(
        ContentPack.id == UUID(pack_id),
        ContentPack.org_id == UUID(org_id),
    ).first()
    if not pack:
        raise HTTPException(404, "Content pack not found")

    # Verify variant exists and belongs to this pack
    variant = db.query(ContentVariant).filter(
        ContentVariant.id == UUID(request.variant_id),
        ContentVariant.content_pack_id == pack.id,
        ContentVariant.channel == request.channel,
    ).first()
    if not variant:
        raise HTTPException(404, "Variant not found for this channel")

    # Upsert lock
    existing = db.query(ContentChannelLock).filter(
        ContentChannelLock.content_pack_id == pack.id,
        ContentChannelLock.channel == request.channel,
    ).first()

    if existing:
        existing.locked_variant_id = variant.id
        existing.locked_at = datetime.utcnow()
        existing.locked_by_user_id = UUID(user_id) if user_id else None
    else:
        lock = ContentChannelLock(
            id=uuid4(),
            content_pack_id=pack.id,
            channel=request.channel,
            locked_variant_id=variant.id,
            locked_by_user_id=UUID(user_id) if user_id else None,
        )
        db.add(lock)

    db.commit()
    return {"status": "locked", "channel": request.channel, "variant_id": str(variant.id)}


@router.get("/packs/{pack_id}/locks")
def get_locks(
    pack_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get all locked variant IDs for a pack."""
    org_id = user.get("org_id", "")
    pack = db.query(ContentPack).filter(
        ContentPack.id == UUID(pack_id),
        ContentPack.org_id == UUID(org_id),
    ).first()
    if not pack:
        raise HTTPException(404, "Content pack not found")

    locks = db.query(ContentChannelLock).filter(
        ContentChannelLock.content_pack_id == pack.id
    ).all()

    return {
        lock.channel: str(lock.locked_variant_id)
        for lock in locks
    }


@router.post("/packs/{pack_id}/regenerate", status_code=202, response_model=AsyncJobResponse)
def regenerate_pack(
    pack_id: str,
    request: RegenerateRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Regenerate non-locked variants in a pack.

    Keeps locked variants untouched, deletes non-locked ones,
    and generates fresh replacements.
    """
    org_id = user.get("org_id", "")
    pack = db.query(ContentPack).filter(
        ContentPack.id == UUID(pack_id),
        ContentPack.org_id == UUID(org_id),
    ).first()
    if not pack:
        raise HTTPException(404, "Content pack not found")

    # Get all locks for this pack
    locks = {
        lock.channel: lock.locked_variant_id
        for lock in db.query(ContentChannelLock).filter(
            ContentChannelLock.content_pack_id == pack.id
        ).all()
    }

    # Determine which channels to regenerate
    all_channels = [ch.get("channel", "") for ch in (pack.channels_json or [])]
    target_channels = request.channels if request.channels else all_channels

    # Delete non-locked variants for target channels
    for channel in target_channels:
        locked_variant_id = locks.get(channel)
        query = db.query(ContentVariant).filter(
            ContentVariant.content_pack_id == pack.id,
            ContentVariant.channel == channel,
        )
        if locked_variant_id:
            query = query.filter(ContentVariant.id != locked_variant_id)
        query.delete(synchronize_session="fetch")

    # Reset pack status
    pack.status = ContentPackStatus.QUEUED
    db.flush()

    # Enqueue regeneration job
    from backend.src.jobs.queue import enqueue

    job_id = enqueue(
        task_name="content_studio_regenerate",
        payload={
            "pack_id": str(pack.id),
            "channels": target_channels,
            "locked_variant_ids": {ch: str(vid) for ch, vid in locks.items()},
        },
        org_id=org_id,
        db=db,
    )

    pack.job_run_id = UUID(job_id)
    db.commit()

    logger.info(
        "CONTENT_STUDIO_REGENERATE | pack={} | channels={} | locks={}",
        pack.id, target_channels, list(locks.keys()),
    )

    return AsyncJobResponse(pack_id=str(pack.id), job_id=job_id, status="queued")


@router.get("/packs/{pack_id}/export/pdf")
def export_pdf(
    pack_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export content pack as PDF production brief."""
    org_id = user.get("org_id", "")
    pack = db.query(ContentPack).filter(
        ContentPack.id == UUID(pack_id),
        ContentPack.org_id == UUID(org_id),
    ).first()
    if not pack:
        raise HTTPException(404, "Content pack not found")

    variants = db.query(ContentVariant).filter(
        ContentVariant.content_pack_id == pack.id
    ).order_by(ContentVariant.channel, ContentVariant.score.desc()).all()

    locks = {
        lock.channel: str(lock.locked_variant_id)
        for lock in db.query(ContentChannelLock).filter(
            ContentChannelLock.content_pack_id == pack.id
        ).all()
    }

    pdf_bytes = _build_pack_pdf(pack, variants, locks)

    filename = f"content_pack_{str(pack.id)[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/packs/{pack_id}/export/xlsx")
def export_xlsx(
    pack_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export content pack as XLSX spreadsheet."""
    org_id = user.get("org_id", "")
    pack = db.query(ContentPack).filter(
        ContentPack.id == UUID(pack_id),
        ContentPack.org_id == UUID(org_id),
    ).first()
    if not pack:
        raise HTTPException(404, "Content pack not found")

    variants = db.query(ContentVariant).filter(
        ContentVariant.content_pack_id == pack.id
    ).order_by(ContentVariant.channel, ContentVariant.variant_index).all()

    xlsx_bytes = _build_pack_xlsx(pack, variants)

    filename = f"content_pack_{str(pack.id)[:8]}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF Builder ──────────────────────────────────────────────────────────────

def _build_pack_pdf(pack, variants, locks) -> bytes:
    """Build production brief PDF."""
    from fpdf import FPDF
    from backend.src.utils.pdf_fonts import setup_pdf_fonts

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    _font = setup_pdf_fonts(pdf)

    pdf.add_page()

    # Title
    pdf.set_font(_font, "B", 18)
    pdf.cell(0, 12, "Content Production Pack", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(_font, "", 10)
    pdf.set_text_color(120, 120, 120)
    settings = pack.input_json or {}
    pdf.cell(0, 6, f"Goal: {pack.goal or 'N/A'} | Audience: {settings.get('target_audience', 'N/A')}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Group variants by channel
    from collections import defaultdict
    by_channel = defaultdict(list)
    for v in variants:
        by_channel[v.channel].append(v)

    for channel_key, channel_variants in by_channel.items():
        spec = CHANNEL_SPECS.get(channel_key)
        display = spec.display_name if spec else channel_key

        pdf.set_font(_font, "B", 14)
        pdf.set_text_color(40, 40, 40)
        pdf.cell(0, 10, display, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        locked_id = locks.get(channel_key)

        for v in channel_variants:
            is_locked = str(v.id) == locked_id
            prefix = "[LOCKED] " if is_locked else ""

            pdf.set_font(_font, "B", 10)
            pdf.set_text_color(60, 60, 60)
            pdf.cell(0, 7, f"{prefix}Variant {v.variant_index} (Score: {v.score:.0f}/100)", new_x="LMARGIN", new_y="NEXT")

            # Output key fields
            output = v.output_json or {}
            pdf.set_font(_font, "", 9)
            pdf.set_text_color(80, 80, 80)

            for field_key, field_val in output.items():
                if isinstance(field_val, list):
                    val_str = ", ".join(str(x) for x in field_val[:10])
                elif isinstance(field_val, dict):
                    val_str = str(field_val)[:200]
                else:
                    val_str = str(field_val)[:300]

                label = field_key.replace("_", " ").title()
                pdf.set_font(_font, "B", 8)
                pdf.cell(40, 5, f"{label}:")
                pdf.set_font(_font, "", 8)
                pdf.multi_cell(0, 5, val_str, new_x="LMARGIN", new_y="NEXT")

            if v.rationale_text:
                pdf.set_font(_font, "I", 8)
                pdf.set_text_color(100, 100, 100)
                pdf.multi_cell(0, 5, f"Rationale: {v.rationale_text[:300]}", new_x="LMARGIN", new_y="NEXT")

            pdf.ln(4)

        pdf.ln(4)

    # Footer
    pdf.set_font(_font, "I", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.cell(0, 5, "Generated by Meta Ops Agent - Content Studio", new_x="LMARGIN", new_y="NEXT")

    return pdf.output()


# ── XLSX Builder ─────────────────────────────────────────────────────────────

def _build_pack_xlsx(pack, variants) -> bytes:
    """Build XLSX with one tab per channel + overview."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Overview sheet
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Content Pack ID", str(pack.id)])
    ws.append(["Goal", pack.goal or ""])
    ws.append(["Language", pack.language or ""])
    ws.append(["Status", pack.status.value if pack.status else ""])
    ws.append(["Created", pack.created_at.isoformat() if pack.created_at else ""])
    ws.append([])
    settings = pack.input_json or {}
    ws.append(["Audience", settings.get("target_audience", "")])
    ws.append(["Tone", ", ".join(settings.get("tone_tags", []))])
    ws.append(["Framework", settings.get("framework_preference", "")])
    ws.append(["Hook Style", settings.get("hook_style_preference", "")])

    # Group by channel
    from collections import defaultdict
    by_channel = defaultdict(list)
    for v in variants:
        by_channel[v.channel].append(v)

    for channel_key, channel_variants in by_channel.items():
        # Create sheet (max 31 chars for sheet name)
        sheet_name = channel_key[:31]
        ws_ch = wb.create_sheet(title=sheet_name)

        # Header row: variant_index + output fields
        if channel_variants:
            sample_output = channel_variants[0].output_json or {}
            headers = ["variant", "score"] + list(sample_output.keys()) + ["rationale"]
            ws_ch.append(headers)

            for v in channel_variants:
                output = v.output_json or {}
                row = [v.variant_index, v.score or 0]
                for key in list(sample_output.keys()):
                    val = output.get(key, "")
                    if isinstance(val, (list, dict)):
                        val = str(val)
                    row.append(str(val)[:500])
                row.append(v.rationale_text or "")
                ws_ch.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.post("/content-studio/packs/{pack_id}/variants/{variant_id}/publish")
def publish_variant_to_meta(
    pack_id: str,
    variant_id: str,
    page_id: str = Query(..., description="Facebook Page ID to publish to"),
    dry_run: bool = Query(True, description="Simulate publish (default: true until App Review)"),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """
    Publish a content variant to a Facebook Page via Meta Graph API.
    Defaults to dry_run=True since pages_manage_posts scope requires App Review.
    """
    from uuid import UUID as _UUID
    import asyncio

    org_id = user.get("org_id", "")
    if not org_id:
        raise HTTPException(400, "Missing org_id")

    # Verify variant belongs to pack
    variant = db.query(ContentVariant).filter(
        ContentVariant.id == _UUID(variant_id),
        ContentVariant.content_pack_id == _UUID(pack_id),
    ).first()
    if not variant:
        raise HTTPException(404, "Variant not found in this pack")

    from backend.src.services.meta_publisher import MetaPublisher
    publisher = MetaPublisher(db, _UUID(org_id))

    # Run async publish in sync context
    loop = asyncio.new_event_loop()
    try:
        result = loop.run_until_complete(
            publisher.publish_variant(
                variant_id=_UUID(variant_id),
                page_id=page_id,
                dry_run=dry_run,
            )
        )
    finally:
        loop.close()

    if result.get("status") == "error":
        raise HTTPException(400, result.get("message", "Publishing failed"))

    return result
