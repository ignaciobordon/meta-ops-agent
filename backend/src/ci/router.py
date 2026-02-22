"""
CI Module — FastAPI Router.

Registered in main.py at /api/ci.
Provides REST endpoints for CI competitor management, item ingestion, search,
detected opportunities listing, and competitive brief exports (PDF/XLSX).
"""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.src.ci.analysis_service import CIAnalysisService
from backend.src.ci.engine import CompetitiveIntelligenceEngine
from backend.src.ci.models import CICanonicalItem, CIRun, CIRunStatus, CIRunType
from backend.src.ci.schemas import (
    CanonicalItemCreate,
    CanonicalItemResponse,
    CompetitorCreate,
    CompetitorResponse,
    CompetitorUpdate,
    DomainResponse,
    SearchRequest,
    SearchResultItem,
    SimilarRequest,
    SourceCreate,
    SourceResponse,
)
from backend.src.database.models import MetaAlert
from backend.src.database.session import get_db
from backend.src.middleware.auth import get_current_user
from src.utils.logging_config import logger

router = APIRouter(tags=["Competitive Intelligence"])


def _get_org_id(user: dict) -> UUID:
    org_id = user.get("org_id", "")
    if not org_id:
        raise HTTPException(400, "Missing org_id in token")
    return UUID(str(org_id)) if not isinstance(org_id, UUID) else org_id


# ── Competitors ───────────────────────────────────────────────────────────────


@router.post("/competitors", response_model=CompetitorResponse, status_code=201)
def create_competitor(
    body: CompetitorCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Register a new competitor."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)

    domains = [{"domain": d.domain, "domain_type": d.domain_type} for d in body.domains]

    competitor = engine.register_competitor(
        org_id=org_id,
        name=body.name,
        website_url=body.website_url,
        logo_url=body.logo_url,
        notes=body.notes,
        domains=domains,
    )
    return _competitor_to_response(competitor)


@router.get("/competitors")
def list_competitors(
    status: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """List competitors for the org with computed stats."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)
    competitors = engine.list_competitors(org_id, status=status, limit=limit, offset=offset)
    return [_competitor_to_response_with_stats(c, db) for c in competitors]


@router.get("/competitors/{competitor_id}", response_model=CompetitorResponse)
def get_competitor(
    competitor_id: UUID,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get a specific competitor."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)
    competitor = engine.get_competitor(org_id, competitor_id)
    if not competitor:
        raise HTTPException(404, "Competitor not found")
    return _competitor_to_response(competitor)


@router.patch("/competitors/{competitor_id}", response_model=CompetitorResponse)
def update_competitor(
    competitor_id: UUID,
    body: CompetitorUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Update a competitor."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)

    updates = body.model_dump(exclude_unset=True)
    competitor = engine.update_competitor(org_id, competitor_id, updates)
    if not competitor:
        raise HTTPException(404, "Competitor not found")
    return _competitor_to_response(competitor)


@router.delete("/competitors/{competitor_id}", status_code=204)
def delete_competitor(
    competitor_id: UUID,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Delete a competitor."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)
    deleted = engine.delete_competitor(org_id, competitor_id)
    if not deleted:
        raise HTTPException(404, "Competitor not found")


@router.post("/competitors/discover")
def discover_competitors(
    query: str = Query(..., description="Industry, niche, or brand name"),
    country: str = Query("", description="Target country code (optional, e.g. AR, US, ES)"),
    limit: int = Query(5, ge=1, le=20),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Use LLM to discover potential competitors for the given niche."""
    _get_org_id(user)  # auth check
    from backend.src.ci.discovery_service import CIDiscoveryService
    service = CIDiscoveryService()
    return service.discover(query=query, country=country, limit=limit)


# ── Sources ───────────────────────────────────────────────────────────────────


@router.post("/sources", response_model=SourceResponse, status_code=201)
def create_source(
    body: SourceCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Create a CI data source."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)
    source = engine.create_source(
        org_id=org_id,
        name=body.name,
        source_type=body.source_type,
        config_json=body.config_json,
    )
    return source


@router.get("/sources", response_model=List[SourceResponse])
def list_sources(
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """List CI sources for the org."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)
    return engine.list_sources(org_id)


# ── Canonical Items ───────────────────────────────────────────────────────────


@router.post("/items", response_model=CanonicalItemResponse, status_code=201)
def upsert_item(
    body: CanonicalItemCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Upsert a canonical CI item."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)

    item = engine.upsert_canonical_item(
        org_id=org_id,
        competitor_id=body.competitor_id,
        item_type=body.item_type,
        external_id=body.external_id or "",
        title=body.title,
        body_text=body.body_text,
        url=body.url,
        image_urls=body.image_urls,
        canonical_json=body.canonical_json,
        raw_json=body.raw_json,
        source_id=body.source_id,
    )

    # Index in vector store
    engine.index_item(item)

    return item


@router.get("/items", response_model=List[CanonicalItemResponse])
def list_items(
    competitor_id: Optional[UUID] = Query(None),
    item_type: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """List canonical items with optional filters."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)
    return engine.list_canonical_items(
        org_id, competitor_id=competitor_id, item_type=item_type,
        limit=limit, offset=offset,
    )


@router.get("/items/{item_id}", response_model=CanonicalItemResponse)
def get_item(
    item_id: UUID,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get a specific canonical item."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)
    item = engine.get_canonical_item(org_id, item_id)
    if not item:
        raise HTTPException(404, "Item not found")
    return item


# ── Search ────────────────────────────────────────────────────────────────────


@router.post("/search")
def search_items(
    body: SearchRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Semantic search across CI items."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)

    results = engine.search_text(
        org_id=org_id,
        query=body.query,
        item_types=body.item_types or None,
        competitor_ids=body.competitor_ids or None,
        n_results=body.n_results,
    )
    return {"results": results, "count": len(results)}


@router.post("/similar")
def find_similar(
    body: SimilarRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Find items similar to a given item."""
    org_id = _get_org_id(user)
    engine = CompetitiveIntelligenceEngine(db)

    results = engine.find_similar(
        org_id=org_id,
        item_id=body.item_id,
        n_results=body.n_results,
    )
    return {"results": results, "count": len(results)}


# ── Helpers ───────────────────────────────────────────────────────────────────


# ── Feed (frontend-format items) ───────────────────────────────────────────


@router.get("/feed")
def feed_items(
    competitor: Optional[str] = Query(None),
    item_type: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    limit: int = Query(20, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """List CI items in the frontend CIItem format used by Radar."""
    org_id = _get_org_id(user)

    query = db.query(CICanonicalItem).filter(CICanonicalItem.org_id == org_id)

    if competitor:
        from backend.src.ci.models import CICompetitor as _Comp
        comp = db.query(_Comp).filter(_Comp.org_id == org_id, _Comp.name.ilike(f"%{competitor}%")).first()
        if comp:
            query = query.filter(CICanonicalItem.competitor_id == comp.id)

    if item_type:
        from backend.src.ci.models import CIItemType
        try:
            query = query.filter(CICanonicalItem.item_type == CIItemType(item_type))
        except ValueError:
            query = query.filter(CICanonicalItem.item_type == item_type)

    if q:
        query = query.filter(
            (CICanonicalItem.title.ilike(f"%{q}%")) | (CICanonicalItem.body_text.ilike(f"%{q}%"))
        )

    items = query.order_by(CICanonicalItem.last_seen_at.desc()).offset(offset).limit(limit).all()

    return [_item_to_feed_format(item) for item in items]


def _item_to_feed_format(item: CICanonicalItem) -> dict:
    """Convert a DB CICanonicalItem to the frontend CIItem shape."""
    canonical = item.canonical_json or {}
    itype = item.item_type.value if hasattr(item.item_type, "value") else str(item.item_type)
    result = {
        "id": str(item.id),
        "competitor": canonical.get("competitor", ""),
        "platform": canonical.get("platform", ""),
        "item_type": itype,
        "headline": item.title or "",
        "body": item.body_text or "",
        "cta": canonical.get("cta", ""),
        "format": canonical.get("format", ""),
        "country": canonical.get("country", ""),
        "price": canonical.get("price"),
        "discount": canonical.get("discount", ""),
        "first_seen": item.first_seen_at.isoformat() if item.first_seen_at else "",
        "last_seen": item.last_seen_at.isoformat() if item.last_seen_at else "",
        "fingerprint": canonical.get("fingerprint", ""),
        "metadata": canonical,
        "has_analysis": item.analysis_json is not None,
    }
    if item.analysis_json and isinstance(item.analysis_json, dict):
        analysis = item.analysis_json.get("analysis", {})
        # Ensure analysis is a dict, not a string
        if isinstance(analysis, str):
            import json as _json
            try:
                analysis = _json.loads(analysis)
            except (ValueError, TypeError):
                analysis = {}
        result["analysis"] = analysis
    return result


# ── Detected Opportunities ─────────────────────────────────────────────────


@router.get("/opportunities")
def list_ci_opportunities(
    type: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """List CI-detected opportunities (from MetaAlert table, entity_type=ci_opportunity)."""
    org_id = _get_org_id(user)

    query = (
        db.query(MetaAlert)
        .filter(
            MetaAlert.org_id == org_id,
            MetaAlert.entity_type == "ci_opportunity",
        )
    )

    if type:
        query = query.filter(MetaAlert.alert_type == type)

    alerts = (
        query
        .order_by(MetaAlert.detected_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    return [
        {
            "id": str(a.id),
            "type": a.alert_type or "ci_new_opportunity",
            "title": (a.message or "")[:200],
            "description": a.message or "",
            "priority_score": _severity_to_score(
                a.severity.value if hasattr(a.severity, "value") else str(a.severity)
            ),
            "confidence_score": (a.payload_json or {}).get("confidence", 0.8),
            "impact_score": _severity_to_score(
                a.severity.value if hasattr(a.severity, "value") else str(a.severity)
            ),
            "evidence_ids": (a.payload_json or {}).get("evidence_ids", []),
            "suggested_actions": (a.payload_json or {}).get("suggested_actions", []),
            "rationale": (a.payload_json or {}).get("rationale", ""),
            "detected_at": a.detected_at.isoformat() if a.detected_at else None,
            "expires_at": None,
            "status": "resolved" if a.resolved_at else "active",
        }
        for a in alerts
    ]


# ── Export ─────────────────────────────────────────────────────────────────


@router.get("/export/pdf")
def export_competitive_brief_pdf(
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export competitive intelligence brief as PDF."""
    org_id = _get_org_id(user)

    competitors = CompetitiveIntelligenceEngine(db).list_competitors(org_id, limit=50)
    opportunities = (
        db.query(MetaAlert)
        .filter(MetaAlert.org_id == org_id, MetaAlert.entity_type == "ci_opportunity")
        .order_by(MetaAlert.detected_at.desc())
        .limit(50)
        .all()
    )
    recent_items = (
        db.query(CICanonicalItem)
        .filter(CICanonicalItem.org_id == org_id)
        .order_by(CICanonicalItem.last_seen_at.desc())
        .limit(100)
        .all()
    )

    pdf_bytes = _build_ci_pdf(competitors, opportunities, recent_items)
    filename = f"competitive_brief_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Analysis Endpoints ──────────────────────────────────────────────────────


@router.post("/items/{item_id}/analyze")
def analyze_item(
    item_id: UUID,
    force_refresh: bool = Query(False),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Trigger LLM strategic analysis for a CI item."""
    org_id = _get_org_id(user)
    service = CIAnalysisService()
    try:
        result = service.analyze_item(db, org_id, item_id, force_refresh=force_refresh)
        return result
    except ValueError as e:
        raise HTTPException(404, str(e))
    except Exception as e:
        logger.error("CI_ANALYSIS_ERROR | item_id={} | error={}", item_id, str(e)[:300])
        raise HTTPException(500, f"Analysis failed: {str(e)[:200]}")


@router.get("/items/{item_id}/analysis")
def get_item_analysis(
    item_id: UUID,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Get cached analysis for a CI item (no LLM call)."""
    org_id = _get_org_id(user)
    service = CIAnalysisService()
    try:
        return service.get_item_analysis(db, org_id, item_id)
    except ValueError as e:
        raise HTTPException(404, str(e))


@router.get("/export/analysis-pdf")
def export_analysis_pdf(
    item_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export analysis as PDF — single item or all analyzed items."""
    org_id = _get_org_id(user)

    if item_id:
        items = (
            db.query(CICanonicalItem)
            .filter(
                CICanonicalItem.org_id == org_id,
                CICanonicalItem.id == item_id,
                CICanonicalItem.analysis_json.isnot(None),
            )
            .all()
        )
    else:
        items = (
            db.query(CICanonicalItem)
            .filter(
                CICanonicalItem.org_id == org_id,
                CICanonicalItem.analysis_json.isnot(None),
            )
            .order_by(CICanonicalItem.last_seen_at.desc())
            .limit(100)
            .all()
        )

    # Filter out items with corrupted analysis_json (e.g. stored as "null" string)
    items = [i for i in items if isinstance(i.analysis_json, dict) and i.analysis_json.get("analysis")]

    if not items:
        raise HTTPException(404, "No analyzed items found")

    # Load brand profile for context header
    from backend.src.database.models import BrandMapProfile
    brand = (
        db.query(BrandMapProfile)
        .filter(BrandMapProfile.org_id == org_id, BrandMapProfile.status == "ready")
        .order_by(BrandMapProfile.last_analyzed_at.desc())
        .first()
    )

    pdf_bytes = _build_analysis_pdf(items, brand)
    filename = f"ci_strategy_report_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/xlsx")
def export_competitive_evidence_xlsx(
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Export competitive evidence as XLSX spreadsheet."""
    org_id = _get_org_id(user)

    competitors = CompetitiveIntelligenceEngine(db).list_competitors(org_id, limit=50)
    items = (
        db.query(CICanonicalItem)
        .filter(CICanonicalItem.org_id == org_id)
        .order_by(CICanonicalItem.last_seen_at.desc())
        .limit(500)
        .all()
    )
    opportunities = (
        db.query(MetaAlert)
        .filter(MetaAlert.org_id == org_id, MetaAlert.entity_type == "ci_opportunity")
        .order_by(MetaAlert.detected_at.desc())
        .limit(100)
        .all()
    )

    xlsx_bytes = _build_ci_xlsx(competitors, items, opportunities)
    filename = f"competitive_evidence_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Helpers ───────────────────────────────────────────────────────────────────


def _severity_to_score(severity: str) -> float:
    return {"critical": 1.0, "high": 0.8, "medium": 0.5, "low": 0.3, "info": 0.1}.get(severity, 0.5)


def _build_ci_pdf(competitors, opportunities, recent_items) -> bytes:
    """Build a competitive intelligence brief PDF."""
    from fpdf import FPDF
    from backend.src.utils.pdf_fonts import setup_pdf_fonts

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    _font = setup_pdf_fonts(pdf)

    pdf.add_page()

    # Title
    pdf.set_font(_font, "B", 18)
    pdf.cell(0, 12, "Competitive Intelligence Brief", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(_font, "", 10)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(
        0, 6,
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | "
        f"{len(competitors)} competitors | {len(opportunities)} opportunities | {len(recent_items)} items",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(8)
    pdf.set_text_color(0, 0, 0)

    # Competitors Section
    pdf.set_font(_font, "B", 14)
    pdf.cell(0, 10, "Competitors Tracked", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    for comp in competitors:
        pdf.set_font(_font, "B", 11)
        name = comp.name if hasattr(comp, "name") else str(comp)
        pdf.cell(0, 7, name, new_x="LMARGIN", new_y="NEXT")
        if hasattr(comp, "website_url") and comp.website_url:
            pdf.set_font(_font, "", 9)
            pdf.set_text_color(80, 80, 80)
            pdf.cell(0, 5, comp.website_url, new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
        domains = getattr(comp, "domains", []) or []
        if domains:
            pdf.set_font(_font, "", 9)
            for d in domains[:5]:
                domain_str = d.domain if hasattr(d, "domain") else str(d)
                pdf.cell(0, 5, f"  - {domain_str}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # Opportunities Section
    if opportunities:
        pdf.add_page()
        pdf.set_font(_font, "B", 14)
        pdf.cell(0, 10, "Detected Opportunities", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        for i, opp in enumerate(opportunities[:20], 1):
            pdf.set_font(_font, "B", 11)
            sev = opp.severity.value if hasattr(opp.severity, "value") else str(opp.severity)
            pdf.cell(0, 7, f"{i}. [{sev.upper()}] {opp.alert_type}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(_font, "", 9)
            message = (opp.message or "")[:300]
            pdf.multi_cell(0, 5, message)
            payload = opp.payload_json or {}
            if payload.get("suggested_actions"):
                pdf.set_font(_font, "I", 9)
                for action in payload["suggested_actions"][:3]:
                    pdf.cell(0, 5, f"  > {action[:100]}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)

    # Recent Items Summary
    if recent_items:
        pdf.add_page()
        pdf.set_font(_font, "B", 14)
        pdf.cell(0, 10, f"Recent Items ({len(recent_items)})", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        for item in recent_items[:30]:
            pdf.set_font(_font, "B", 10)
            itype = item.item_type.value if hasattr(item.item_type, "value") else str(item.item_type)
            title = (item.title or "(untitled)")[:80]
            pdf.cell(0, 6, f"[{itype}] {title}", new_x="LMARGIN", new_y="NEXT")
            pdf.set_font(_font, "", 8)
            seen = item.last_seen_at.strftime("%Y-%m-%d") if item.last_seen_at else "?"
            pdf.cell(0, 4, f"  Last seen: {seen}", new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)

    return pdf.output()


# _safe_text is now a thin wrapper around the shared utility.
# It captures the font family from the enclosing _build_analysis_pdf scope.
_analysis_font_family = "Helvetica"


def _safe_text(text: str) -> str:
    """Sanitize text for fpdf — delegates to shared safe_text utility."""
    if not text:
        return ""
    import re as _re
    text = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    from backend.src.utils.pdf_fonts import safe_text
    return safe_text(text, _analysis_font_family)


def _build_analysis_pdf(items, brand_profile=None) -> bytes:
    """Build a strategic analysis PDF from analyzed CI items."""
    from fpdf import FPDF
    from backend.src.utils.pdf_fonts import setup_pdf_fonts

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    global _analysis_font_family
    _font = setup_pdf_fonts(pdf)
    _analysis_font_family = _font

    def _pdf_write(text: str, *, bold: bool = False, size: int = 9,
                   fill_color=None, is_cell: bool = False):
        """Safe wrapper for pdf text output — catches rendering errors."""
        try:
            pdf.set_font(_font, "B" if bold else "", size)
            if fill_color:
                pdf.set_fill_color(*fill_color)
            safe = _safe_text(text)
            if is_cell:
                pdf.cell(0, 6 if bold else 5, safe, new_x="LMARGIN", new_y="NEXT",
                         fill=bool(fill_color))
            else:
                pdf.multi_cell(0, 5, safe)
        except Exception:
            pass  # Skip text that can't be rendered

    pdf.add_page()

    # ── Title Page ──
    pdf.set_font(_font, "B", 20)
    pdf.cell(0, 14, "Competitive Intelligence", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font(_font, "B", 16)
    pdf.cell(0, 10, "Strategic Analysis Report", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    pdf.set_font(_font, "", 10)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(
        0, 6,
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | {len(items)} items analyzed",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.set_text_color(0, 0, 0)
    pdf.ln(6)

    # ── Brand Context Summary ──
    if brand_profile and brand_profile.structured_json:
        sj = brand_profile.structured_json
        pdf.set_font(_font, "B", 13)
        pdf.set_fill_color(240, 245, 255)
        pdf.cell(0, 8, "Brand Context", new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.ln(2)
        pdf.set_font(_font, "", 9)
        ci = sj.get("core_identity", {})
        if ci.get("mission"):
            pdf.multi_cell(0, 5, _safe_text(f"Mission: {ci['mission'][:200]}"))
        if sj.get("tone_voice"):
            tv = sj["tone_voice"]
            if isinstance(tv, dict):
                tv_str = ", ".join(f"{k}: {v}" for k, v in list(tv.items())[:4])
            else:
                tv_str = str(tv)[:200]
            pdf.multi_cell(0, 5, _safe_text(f"Tone & Voice: {tv_str}"))
        diff = sj.get("differentiation", {})
        if isinstance(diff, dict) and diff.get("usp"):
            pdf.multi_cell(0, 5, _safe_text(f"USP: {diff['usp'][:200]}"))
        pdf.ln(6)

    # ── Per-item Analysis Pages ──
    for idx, item in enumerate(items):
        if idx > 0:
            pdf.add_page()

        analysis_data = item.analysis_json if isinstance(item.analysis_json, dict) else {}
        analysis = analysis_data.get("analysis", {})
        if isinstance(analysis, str):
            import json as _json
            try:
                analysis = _json.loads(analysis)
            except (ValueError, TypeError):
                analysis = {}
        canonical = item.canonical_json or {}
        itype = item.item_type.value if hasattr(item.item_type, "value") else str(item.item_type)

        # Item header
        pdf.set_font(_font, "B", 13)
        competitor_name = canonical.get("competitor", "Unknown")
        pdf.set_fill_color(245, 245, 250)
        pdf.cell(0, 8, f"{competitor_name} | {itype.upper()}", new_x="LMARGIN", new_y="NEXT", fill=True)
        pdf.ln(2)

        # Original content
        _pdf_write("Original Content", bold=True, size=10, is_cell=True)
        if item.title:
            _pdf_write(f"Headline: {item.title[:300]}")
        if item.body_text:
            _pdf_write(f"Body: {item.body_text[:500]}")
        if canonical.get("cta"):
            _pdf_write(f"CTA: {canonical['cta'][:100]}", is_cell=True)
        if canonical.get("offers"):
            _pdf_write(f"Offers: {'; '.join(canonical['offers'][:5])}")
        pdf.ln(4)

        # Competitor Strategy
        if analysis.get("competitor_strategy"):
            _pdf_write("Competitor Strategy", bold=True, size=10,
                       fill_color=(255, 248, 240), is_cell=True)
            _pdf_write(str(analysis["competitor_strategy"])[:600])
            pdf.ln(2)

        # Messaging Angles
        if analysis.get("messaging_angles"):
            _pdf_write("Messaging Angles", bold=True, size=10, is_cell=True)
            for angle in analysis["messaging_angles"][:6]:
                _pdf_write(f"  - {str(angle)[:150]}", is_cell=True)
            pdf.ln(2)

        # Brand Comparison
        if analysis.get("brand_comparison"):
            _pdf_write("Brand Comparison", bold=True, size=10,
                       fill_color=(240, 255, 245), is_cell=True)
            _pdf_write(str(analysis["brand_comparison"])[:600])
            pdf.ln(2)

        # Recommendations
        if analysis.get("recommendations"):
            _pdf_write("Actionable Recommendations", bold=True, size=10,
                       fill_color=(240, 248, 255), is_cell=True)
            for i, rec in enumerate(analysis["recommendations"][:5], 1):
                _pdf_write(f"  {i}. {str(rec)[:200]}")
            pdf.ln(2)

        # Ad Copy Suggestions
        if analysis.get("ad_copy_suggestions"):
            _pdf_write("Suggested Ad Copy Variations", bold=True, size=10,
                       fill_color=(250, 240, 255), is_cell=True)
            for i, copy in enumerate(analysis["ad_copy_suggestions"][:4], 1):
                _pdf_write(f"  {i}. {str(copy)[:300]}")
            pdf.ln(2)

        # Opportunities
        if analysis.get("opportunities"):
            _pdf_write("Strategic Opportunities", bold=True, size=10, is_cell=True)
            for opp in analysis["opportunities"][:4]:
                _pdf_write(f"  - {str(opp)[:200]}")
            pdf.ln(2)

        # Threat Level
        threat = analysis.get("threat_level", "medium")
        pdf.set_font(_font, "B", 10)
        threat_colors = {"low": (76, 175, 80), "medium": (255, 152, 0), "high": (244, 67, 54)}
        color = threat_colors.get(threat, (255, 152, 0))
        pdf.set_text_color(*color)
        pdf.cell(0, 6, f"Threat Level: {threat.upper()}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)

        # Model / generated info
        pdf.ln(3)
        pdf.set_font(_font, "I", 8)
        pdf.set_text_color(150, 150, 150)
        gen_at = analysis_data.get("generated_at", "")[:19]
        model = analysis_data.get("model", "")
        pdf.cell(0, 4, f"Generated: {gen_at} | Model: {model}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)

    return pdf.output()


def _build_ci_xlsx(competitors, items, opportunities) -> bytes:
    """Build a competitive evidence XLSX spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")

    # Sheet 1: Competitors
    ws = wb.active
    ws.title = "Competitors"
    headers = ["Name", "Website", "Status", "Domains", "Created"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, comp in enumerate(competitors, 2):
        ws.cell(row=row, column=1, value=comp.name)
        ws.cell(row=row, column=2, value=comp.website_url or "")
        status = comp.status.value if hasattr(comp.status, "value") else str(comp.status)
        ws.cell(row=row, column=3, value=status)
        domains = ", ".join(d.domain for d in (comp.domains or []))
        ws.cell(row=row, column=4, value=domains)
        ws.cell(row=row, column=5, value=comp.created_at.strftime("%Y-%m-%d") if comp.created_at else "")

    # Sheet 2: Items (Evidence)
    ws2 = wb.create_sheet("Items")
    headers2 = ["Type", "Title", "Body", "URL", "Competitor ID", "First Seen", "Last Seen"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, item in enumerate(items, 2):
        itype = item.item_type.value if hasattr(item.item_type, "value") else str(item.item_type)
        ws2.cell(row=row, column=1, value=itype)
        ws2.cell(row=row, column=2, value=(item.title or "")[:200])
        ws2.cell(row=row, column=3, value=(item.body_text or "")[:500])
        ws2.cell(row=row, column=4, value=(item.url or "")[:300])
        ws2.cell(row=row, column=5, value=str(item.competitor_id))
        ws2.cell(row=row, column=6, value=item.first_seen_at.strftime("%Y-%m-%d") if item.first_seen_at else "")
        ws2.cell(row=row, column=7, value=item.last_seen_at.strftime("%Y-%m-%d") if item.last_seen_at else "")

    # Sheet 3: Opportunities
    ws3 = wb.create_sheet("Opportunities")
    headers3 = ["Type", "Severity", "Message", "Rationale", "Actions", "Detected"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, opp in enumerate(opportunities, 2):
        ws3.cell(row=row, column=1, value=opp.alert_type or "")
        sev = opp.severity.value if hasattr(opp.severity, "value") else str(opp.severity)
        ws3.cell(row=row, column=2, value=sev)
        ws3.cell(row=row, column=3, value=(opp.message or "")[:500])
        payload = opp.payload_json or {}
        ws3.cell(row=row, column=4, value=payload.get("rationale", "")[:300])
        ws3.cell(row=row, column=5, value="; ".join(payload.get("suggested_actions", []))[:300])
        ws3.cell(row=row, column=6, value=opp.detected_at.strftime("%Y-%m-%d %H:%M") if opp.detected_at else "")

    # Auto-width
    for ws_sheet in [ws, ws2, ws3]:
        for col_cells in ws_sheet.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws_sheet.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _competitor_to_response(c) -> CompetitorResponse:
    """Convert ORM CICompetitor to response schema."""
    return CompetitorResponse(
        id=c.id,
        org_id=c.org_id,
        name=c.name,
        website_url=c.website_url,
        logo_url=c.logo_url,
        notes=c.notes,
        status=c.status.value if hasattr(c.status, "value") else str(c.status),
        meta_json=c.meta_json or {},
        domains=[
            DomainResponse(
                id=d.id,
                domain=d.domain,
                domain_type=d.domain_type.value if hasattr(d.domain_type, "value") else str(d.domain_type),
                verified=d.verified,
                created_at=d.created_at,
            )
            for d in (c.domains or [])
        ],
        created_at=c.created_at,
        updated_at=c.updated_at,
    )


def _competitor_to_response_with_stats(c, db: Session) -> dict:
    """Convert ORM CICompetitor to dict with computed stats for the frontend."""
    from datetime import datetime, timedelta
    from sqlalchemy import func

    now = datetime.utcnow()
    thirty_days_ago = now - timedelta(days=30)

    total_ads = db.query(func.count(CICanonicalItem.id)).filter(
        CICanonicalItem.competitor_id == c.id,
    ).scalar() or 0

    active_ads = db.query(func.count(CICanonicalItem.id)).filter(
        CICanonicalItem.competitor_id == c.id,
        CICanonicalItem.last_seen_at >= thirty_days_ago,
    ).scalar() or 0

    last_item = db.query(CICanonicalItem.last_seen_at).filter(
        CICanonicalItem.competitor_id == c.id,
    ).order_by(CICanonicalItem.last_seen_at.desc()).first()

    last_seen = last_item[0].isoformat() if last_item and last_item[0] else None

    # Infer platform from domains or default
    platform = "meta"
    if c.domains:
        for d in c.domains:
            dt = d.domain_type.value if hasattr(d.domain_type, "value") else str(d.domain_type)
            if "meta" in dt or "ad_library" in dt:
                platform = "meta"
                break
            elif "google" in dt:
                platform = "google"
                break
            elif "tiktok" in dt:
                platform = "tiktok"
                break

    return {
        "id": str(c.id),
        "name": c.name,
        "website_url": c.website_url,
        "logo_url": c.logo_url,
        "notes": c.notes,
        "status": c.status.value if hasattr(c.status, "value") else str(c.status),
        "platform": platform,
        "total_ads": total_ads,
        "active_ads": active_ads,
        "last_seen": last_seen,
        "domains": [
            {
                "id": str(d.id),
                "domain": d.domain,
                "domain_type": d.domain_type.value if hasattr(d.domain_type, "value") else str(d.domain_type),
            }
            for d in (c.domains or [])
        ],
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }
