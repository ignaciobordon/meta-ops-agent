"""Tests for Data Room export job lifecycle (Sprint 13)."""
import pytest
from unittest.mock import MagicMock, patch
from uuid import uuid4
from datetime import datetime


class TestDataRoomExportJob:

    def test_data_export_model_exists(self):
        """DataExport model exists with expected columns."""
        from backend.src.database.models import DataExport
        assert DataExport.__tablename__ == "data_exports"
        columns = [c.name for c in DataExport.__table__.columns]
        assert "org_id" in columns
        assert "status" in columns
        assert "file_path" in columns
        assert "rows_exported" in columns
        assert "params_json" in columns

    def test_export_service_build_xlsx_returns_bytes(self):
        """build_xlsx should return (bytes, int)."""
        from backend.src.services.data_room_export_service import DataRoomExportService
        db = MagicMock()
        org_id = uuid4()
        svc = DataRoomExportService(db, org_id)

        # Mock all queries to return empty
        mock_query = MagicMock()
        db.query.return_value = mock_query
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.limit.return_value = mock_query
        mock_query.all.return_value = []
        mock_query.first.return_value = None

        xlsx_bytes, total_rows = svc.build_xlsx({"datasets": ["job_runs"]})
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0
        assert isinstance(total_rows, int)

    def test_export_service_build_xlsx_empty_produces_valid_xlsx(self):
        """Empty export produces a valid XLSX file."""
        import io
        from openpyxl import load_workbook
        from backend.src.services.data_room_export_service import DataRoomExportService

        db = MagicMock()
        svc = DataRoomExportService(db, uuid4())
        mock_query = MagicMock()
        db.query.return_value = mock_query
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value = mock_query
        mock_query.limit.return_value = mock_query
        mock_query.all.return_value = []
        mock_query.first.return_value = None

        xlsx_bytes, _ = svc.build_xlsx({"datasets": ["alerts"]})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.sheetnames) >= 1

    def test_job_type_registered(self):
        """data_room_export should be in QUEUE_ROUTING."""
        from backend.src.jobs.queue import QUEUE_ROUTING
        assert "data_room_export" in QUEUE_ROUTING

    def test_job_timeout_set(self):
        """data_room_export should have timeout configured."""
        from backend.src.jobs.task_runner import _JOB_TIMEOUT
        assert "data_room_export" in _JOB_TIMEOUT
        assert _JOB_TIMEOUT["data_room_export"] >= 120
