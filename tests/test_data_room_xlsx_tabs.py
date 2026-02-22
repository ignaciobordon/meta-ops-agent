"""Tests for Data Room XLSX tab generation (Sprint 13)."""
import io
import pytest
from unittest.mock import MagicMock
from uuid import uuid4
from openpyxl import load_workbook


def _make_empty_service():
    from backend.src.services.data_room_export_service import DataRoomExportService
    db = MagicMock()
    svc = DataRoomExportService(db, uuid4())
    mock_query = MagicMock()
    db.query.return_value = mock_query
    mock_query.filter.return_value = mock_query
    mock_query.join.return_value = mock_query
    mock_query.order_by.return_value = mock_query
    mock_query.limit.return_value = mock_query
    mock_query.all.return_value = []
    mock_query.first.return_value = None
    return svc


class TestDataRoomXlsxTabs:

    def test_xlsx_creates_requested_tabs(self):
        """Only requested datasets should appear as sheets."""
        svc = _make_empty_service()
        xlsx_bytes, _ = svc.build_xlsx({"datasets": ["alerts", "job_runs"]})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.sheetnames) >= 2

    def test_xlsx_with_all_datasets(self):
        """Requesting all datasets creates at least 10 sheets."""
        svc = _make_empty_service()
        all_keys = [
            "flywheel_runs", "flywheel_steps", "decision_queue",
            "decision_outcomes", "decision_rankings", "job_runs",
            "alerts", "opportunities", "creatives", "meta_insights_daily",
        ]
        xlsx_bytes, _ = svc.build_xlsx({"datasets": all_keys})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.sheetnames) >= 10

    def test_xlsx_row_limit_param(self):
        """row_limit parameter should be accepted."""
        svc = _make_empty_service()
        xlsx_bytes, rows = svc.build_xlsx({"datasets": ["job_runs"], "row_limit": 5})
        assert isinstance(rows, int)

    def test_xlsx_date_filter_param(self):
        """date_from and date_to params should be accepted."""
        svc = _make_empty_service()
        xlsx_bytes, _ = svc.build_xlsx({
            "datasets": ["alerts"],
            "date_from": "2024-01-01",
            "date_to": "2024-12-31",
        })
        assert isinstance(xlsx_bytes, bytes)

    def test_xlsx_headers_present(self):
        """Each sheet should have header row."""
        svc = _make_empty_service()
        xlsx_bytes, _ = svc.build_xlsx({"datasets": ["alerts"]})
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.worksheets[0]
        # First row should have header values
        headers = [cell.value for cell in ws[1] if cell.value]
        assert len(headers) > 0
